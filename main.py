#!/usr/bin/env python3
import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog, messagebox, LabelFrame
import sqlite3
import subprocess
import threading
import datetime
import hashlib
import os
import platform
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import PageBreak

MODULES = [
    ("Access Control", "access_control.sh"),
    ("Package Management", "package_mgmt.sh"),
    ("Filesystem", "filesystem.sh"),
    ("Services", "services.sh"),
    ("System Maintenance", "system_maintenance.sh"),
    ("Firewall", "firewall.sh"),
    ("Network", "network.sh"),
    ("User Accounts", "user_accounts.sh"),
    ("Logging and Auditing", "logging_auditing.sh"),
]

ROLLBACK_SCRIPTS = {
    "Access Control": "access_control_rollback.sh",
    "Package Management": "package_mgmt_rollback.sh",
    "Filesystem": "filesystem_rollback.sh",
    "Services": "services_rollback.sh",
    "System Maintenance": "system_maintenance_rollback.sh",
    "Firewall": "firewall_rollback.sh",
    "Network": "network_rollback.sh",
    "User Accounts": "user_accounts_rollback.sh",
    "Logging and Auditing": "logging_auditing_rollback.sh",
}

BG_COLOR = "#f0f8ff"
HEADER_COLOR = "#1e88e5"
SIDEBAR_COLOR = "#bbdefb"
BUTTON_COLOR = "#2196f3"
BUTTON_HOVER = "#1976d2"
TEXT_COLOR = "#212121"
COLOR_PASS = "#4caf50"
COLOR_FAIL = "#f44336"
COLOR_WARN = "#ff9800"
COLOR_INFO = "#2196f3"
COLOR_FIXED = "#673ab7"
COLOR_MANUAL = "#ff9800"
COLOR_RUNNING = "#9c27b0"
COLOR_NORMAL = "#212121"


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
DB_FILE = os.path.join(PARENT_DIR, "hardening.db")

def run_command_stream(cmd, line_callback, done_callback=None, progress_callback=None):
    try:
        proc = subprocess.Popen(cmd,
                                stdout=subprocess.PIPE,
                                stderr=subprocess.STDOUT,
                                bufsize=1,
                                universal_newlines=True,
                                text=True)
    except Exception as e:
        line_callback(f"[ERROR] Failed to start: {e}")
        if done_callback:
            done_callback(-1)
        return

    lines_processed = 0
    
    try:
        for line in proc.stdout:
            if not line:
                continue
            stripped_line = line.rstrip("\n")
            line_callback(stripped_line)
            lines_processed += 1
            if progress_callback:
                progress_callback(lines_processed)
    except Exception as e:
        line_callback(f"[ERROR] Exception while reading output: {e}")
    proc.wait()
    if done_callback:
        done_callback(proc.returncode)

class PDFReportGenerator:
    def __init__(self, db_conn):
        self.db = db_conn
    
    def get_system_info(self):
        try:
            with open('/etc/os-release', 'r') as f:
                os_info = {}
                for line in f:
                    if '=' in line:
                        key, value = line.strip().split('=', 1)
                        os_info[key] = value.strip('"')
            
            distribution = os_info.get('NAME', 'Unknown')
            version = os_info.get('VERSION', 'Unknown')
            
            return {
                'os_name': f"{distribution} {version}",
                'architecture': platform.machine(),
                'kernel': platform.release(),
                'hostname': platform.node(),
                'distribution_id': os_info.get('ID', 'Unknown').upper()
            }
        except:
            return {
                'os_name': platform.system(),
                'architecture': platform.machine(),
                'kernel': platform.release(),
                'hostname': platform.node(),
                'distribution_id': platform.system()
            }
    
    def calculate_compliance_stats(self, module_name=None):
        cursor = self.db.cursor()
        
        # Get only latest results for accurate stats
        if module_name:
            query = """
                SELECT s1.status, COUNT(*) as count 
                FROM scan_results s1
                INNER JOIN (
                    SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                    FROM scan_results 
                    WHERE module_name=?
                    GROUP BY policy_id, module_name
                ) s2 ON s1.policy_id = s2.policy_id 
                    AND s1.module_name = s2.module_name 
                    AND s1.scan_timestamp = s2.max_timestamp
                WHERE s1.module_name=?
                GROUP BY s1.status
            """
            cursor.execute(query, (module_name, module_name))
        else:
            query = """
                SELECT s1.status, COUNT(*) as count 
                FROM scan_results s1
                INNER JOIN (
                    SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                    FROM scan_results 
                    GROUP BY policy_id, module_name
                ) s2 ON s1.policy_id = s2.policy_id 
                    AND s1.module_name = s2.module_name 
                    AND s1.scan_timestamp = s2.max_timestamp
                GROUP BY s1.status
            """
            cursor.execute(query)
        
        stats = cursor.fetchall()
        
        total = 0
        passed = 0
        failed = 0
        manual = 0
        warning = 0
        
        for stat in stats:
            count = stat['count']
            status = stat['status']
            total += count
            
            if status == "PASS":
                passed += count
            elif status == "FAIL":
                failed += count
            elif status == "MANUAL":
                manual += count
            elif status in ["WARN", "WARNING"]:
                warning += count
        
        if (passed + failed) > 0:
            compliance_pct = (passed / (passed + failed)) * 100
        else:
            compliance_pct = 0
        
        if compliance_pct >= 90:
            risk_level = "LOW"
        elif compliance_pct >= 70:
            risk_level = "MEDIUM"
        else:
            risk_level = "HIGH"
        
        return {
            'total_rules': total,
            'passed': passed,
            'failed': failed,
            'manual': manual,
            'warnings': warning,
            'compliance_pct': round(compliance_pct, 1),
            'risk_level': risk_level
        }
    
    def generate_report(self, module_name=None, output_path=None):
        if output_path is None:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            if module_name:
                output_path = f"hardening_report_{module_name}_{timestamp}.pdf"
            else:
                output_path = f"hardening_report_all_{timestamp}.pdf"
        
        doc = SimpleDocTemplate(output_path, pagesize=A4,
                               leftMargin=0.5*inch, rightMargin=0.5*inch,
                               topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=15,
            textColor=colors.HexColor('#1e88e5')
        )
        
        section_style = ParagraphStyle(
            'Section',
            parent=styles['Heading3'],
            fontSize=11,
            spaceBefore=10,
            spaceAfter=6,
            textColor=colors.HexColor('#37474f'),
            fontName='Helvetica-Bold'
        )
        
        normal_small = ParagraphStyle(
            'NormalSmall',
            parent=styles['Normal'],
            fontSize=8,
            leading=10
        )
        
        table_cell_style = ParagraphStyle(
            'TableCell',
            parent=styles['Normal'],
            fontSize=8,
            leading=9,
            wordWrap='CJK'
        )
        
        system_info = self.get_system_info()
        stats = self.calculate_compliance_stats(module_name)
        
        story.append(Paragraph("Linux Hardening Compliance Report", title_style))
        story.append(Spacer(1, 5))
        
        meta_data = [
            ["Report Generated:", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["Operating System:", system_info['os_name']],
            ["Architecture:", system_info['architecture']],
            ["Kernel Version:", system_info['kernel']],
            ["Distribution ID:", system_info['distribution_id']],
            ["Module:", module_name if module_name else 'All Modules'],
            ["Report ID:", f"HARDEN-{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"]
        ]
        
        meta_table_data = []
        for label, value in meta_data:
            meta_table_data.append([
                Paragraph(f"<b>{label}</b>", table_cell_style),
                Paragraph(value, table_cell_style)
            ])
        
        meta_table = Table(meta_table_data, colWidths=[1.2*inch, 3.8*inch])
        meta_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
        ]))
        
        story.append(meta_table)
        story.append(Spacer(1, 15))
        
        story.append(Paragraph("Executive Summary", section_style))
        
        exec_data = [
            ["Total Rules Scanned:", str(stats['total_rules'])],
            ["Rules Passed:", str(stats['passed'])],
            ["Rules Failed:", str(stats['failed'])],
            ["Warnings:", str(stats['warnings'])],
            ["Overall Compliance:", f"{stats['compliance_pct']}%"],
            ["Risk Level:", f"<font color=\"{'#4caf50' if stats['risk_level'] == 'LOW' else '#ff9800' if stats['risk_level'] == 'MEDIUM' else '#f44336'}\">{stats['risk_level']}</font>"]
        ]
        
        exec_table_data = []
        for label, value in exec_data:
            exec_table_data.append([
                Paragraph(f"<b>{label}</b>", table_cell_style),
                Paragraph(value, table_cell_style)
            ])
        
        exec_table = Table(exec_table_data, colWidths=[1.5*inch, 1*inch])
        exec_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ]))
        
        story.append(exec_table)
        story.append(Spacer(1, 15))
        
        story.append(Paragraph("Detailed Results", section_style))
        story.append(Spacer(1, 5))
        
        cursor = self.db.cursor()
        
        if module_name:
            # Get only latest results for each policy
            cursor.execute("""
                SELECT s1.policy_id, s1.policy_name, s1.expected_value, s1.current_value, s1.status 
                FROM scan_results s1
                INNER JOIN (
                    SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                    FROM scan_results 
                    WHERE module_name=?
                    GROUP BY policy_id, module_name
                ) s2 ON s1.policy_id = s2.policy_id 
                    AND s1.module_name = s2.module_name 
                    AND s1.scan_timestamp = s2.max_timestamp
                WHERE s1.module_name=?
                ORDER BY s1.policy_id
            """, (module_name, module_name))
        else:
            # Get only latest results for each policy across all modules
            cursor.execute("""
                SELECT s1.module_name, s1.policy_id, s1.policy_name, s1.expected_value, s1.current_value, s1.status 
                FROM scan_results s1
                INNER JOIN (
                    SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                    FROM scan_results 
                    GROUP BY policy_id, module_name
                ) s2 ON s1.policy_id = s2.policy_id 
                    AND s1.module_name = s2.module_name 
                    AND s1.scan_timestamp = s2.max_timestamp
                ORDER BY s1.module_name, s1.policy_id
            """)
        
        rows = cursor.fetchall()
        
        if not rows:
            story.append(Paragraph("No scan results available.", normal_small))
        else:
            if module_name:
                table_data = []
                
                header_row = [
                    Paragraph("<b>ID</b>", table_cell_style),
                    Paragraph("<b>Policy Name</b>", table_cell_style),
                    Paragraph("<b>Expected Value</b>", table_cell_style),
                    Paragraph("<b>Current Value</b>", table_cell_style),
                    Paragraph("<b>Status</b>", table_cell_style)
                ]
                table_data.append(header_row)
                
                for row in rows:
                    policy_name = str(row['policy_name'])
                    if len(policy_name) > 60:
                        policy_name = policy_name[:57] + "..."
                    
                    expected = str(row['expected_value'] or "")
                    if len(expected) > 30:
                        expected = expected[:27] + "..."
                    
                    current = str(row['current_value'] or "")
                    if len(current) > 30:
                        current = current[:27] + "..."
                    
                    table_data.append([
                        Paragraph(str(row['policy_id']), table_cell_style),
                        Paragraph(policy_name, table_cell_style),
                        Paragraph(expected, table_cell_style),
                        Paragraph(current, table_cell_style),
                        Paragraph(str(row['status']), table_cell_style)
                    ])
                
                col_widths = [0.4*inch, 2.5*inch, 1.2*inch, 1.2*inch, 0.6*inch]
                
            else:
                table_data = []
                
                header_row = [
                    Paragraph("<b>Module</b>", table_cell_style),
                    Paragraph("<b>ID</b>", table_cell_style),
                    Paragraph("<b>Policy Name</b>", table_cell_style),
                    Paragraph("<b>Status</b>", table_cell_style),
                    Paragraph("<b>Expected</b>", table_cell_style),
                    Paragraph("<b>Current</b>", table_cell_style)
                ]
                table_data.append(header_row)
                
                for row in rows:
                    module = str(row['module_name'])
                    policy_name = str(row['policy_name'])
                    if len(policy_name) > 40:
                        policy_name = policy_name[:37] + "..."
                    
                    expected = str(row['expected_value'] or "")
                    if len(expected) > 15:
                        expected = expected[:12] + "..."
                    
                    current = str(row['current_value'] or "")
                    if len(current) > 15:
                        current = current[:12] + "..."
                    
                    table_data.append([
                        Paragraph(module, table_cell_style),
                        Paragraph(str(row['policy_id']), table_cell_style),
                        Paragraph(policy_name, table_cell_style),
                        Paragraph(str(row['status']), table_cell_style),
                        Paragraph(expected, table_cell_style),
                        Paragraph(current, table_cell_style)
                    ])
                
                col_widths = [0.7*inch, 0.4*inch, 2.2*inch, 0.6*inch, 0.8*inch, 0.8*inch]
            
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e88e5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
                ('TOPPADDING', (0, 1), (-1, -1), 3),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ])
            
            status_col_index = 4 if module_name else 3
            
            for i in range(1, len(table_data)):
                status_cell = table_data[i][status_col_index]
                status_text = status_cell.getPlainText() if hasattr(status_cell, 'getPlainText') else str(status_cell)
                
                if "PASS" in status_text:
                    bg_color = colors.HexColor('#4caf50')
                elif "FAIL" in status_text:
                    bg_color = colors.HexColor('#f44336')
                elif "MANUAL" in status_text:
                    bg_color = colors.HexColor('#ff9800')
                elif "WARN" in status_text or "WARNING" in status_text:
                    bg_color = colors.HexColor('#ffc107')
                else:
                    bg_color = colors.HexColor('#9e9e9e')
                
                style.add('BACKGROUND', (status_col_index, i), (status_col_index, i), bg_color)
                style.add('TEXTCOLOR', (status_col_index, i), (status_col_index, i), colors.white)
            
            table.setStyle(style)
            story.append(table)
            
            if len(rows) > 20:
                story.append(PageBreak())
            
            story.append(Spacer(1, 10))
            
            story.append(Paragraph("Compliance Summary", section_style))
            story.append(Spacer(1, 5))
            
            if module_name:
                summary_data = [
                    ["Status", "Count", "Percentage"],
                    ["Passed", str(stats['passed']), f"{stats['compliance_pct']}%"],
                    ["Failed", str(stats['failed']), f"{100 - stats['compliance_pct']}%"],
                    ["Warnings", str(stats['warnings']), f"{(stats['warnings']/stats['total_rules']*100):.1f}%" if stats['total_rules'] > 0 else "0%"],
                    ["Total", str(stats['total_rules']), "100%"]
                ]
            else:
                # Get module stats from latest results only
                cursor.execute("""
                    SELECT s1.module_name, 
                           COUNT(*) as total,
                           SUM(CASE WHEN s1.status='PASS' THEN 1 ELSE 0 END) as passed,
                           SUM(CASE WHEN s1.status='FAIL' THEN 1 ELSE 0 END) as failed
                    FROM scan_results s1
                    INNER JOIN (
                        SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                        FROM scan_results 
                        GROUP BY policy_id, module_name
                    ) s2 ON s1.policy_id = s2.policy_id 
                        AND s1.module_name = s2.module_name 
                        AND s1.scan_timestamp = s2.max_timestamp
                    GROUP BY s1.module_name
                    ORDER BY s1.module_name
                """)
                module_stats = cursor.fetchall()
                
                summary_data = [["Module", "Total", "Passed", "Failed", "Compliance"]]
                for stat in module_stats:
                    total = stat['total']
                    passed = stat['passed'] or 0
                    failed = stat['failed'] or 0
                    compliance = (passed / total * 100) if total > 0 else 0
                    
                    summary_data.append([
                        stat['module_name'],
                        str(total),
                        str(passed),
                        str(failed),
                        f"{compliance:.1f}%"
                    ])
                
                summary_data.append([
                    "TOTAL",
                    str(stats['total_rules']),
                    str(stats['passed']),
                    str(stats['failed']),
                    f"{stats['compliance_pct']}%"
                ])
            
            summary_table_data = []
            for i, row in enumerate(summary_data):
                row_data = []
                for j, cell in enumerate(row):
                    if i == 0:
                        row_data.append(Paragraph(f"<b>{cell}</b>", table_cell_style))
                    else:
                        row_data.append(Paragraph(str(cell), table_cell_style))
                summary_table_data.append(row_data)
            
            summary_table = Table(summary_table_data, colWidths=[1.5*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.8*inch])
            summary_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#607d8b')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ])
            
            if not module_name:
                summary_style.add('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e3f2fd'))
                summary_style.add('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')
            
            summary_table.setStyle(summary_style)
            story.append(summary_table)
        
        story.append(Spacer(1, 15))
        story.append(Paragraph("="*80, normal_small))
        
        report_data = str(rows) + str(datetime.datetime.now()) + str(stats)
        report_hash = hashlib.sha256(report_data.encode()).hexdigest()
        
        integrity_text = f"""
        <b>Integrity Verification:</b><br/>
        <font size="7">Document Hash: {report_hash[:32]}...<br/>
        Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
        System: {system_info['hostname']}<br/>
        To verify: Compare this hash with stored hash in database.</font>
        """
        story.append(Paragraph(integrity_text, normal_small))
        
        doc.build(story)
        
        self.store_report_hash(output_path, report_hash, module_name)
        
        return output_path, report_hash
        
    def store_report_hash(self, filename, report_hash, module_name):
        try:
            cursor = self.db.cursor()
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS report_hashes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    filename TEXT NOT NULL,
                    hash TEXT NOT NULL,
                    module_name TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            cursor.execute('''
                INSERT INTO report_hashes (filename, hash, module_name) 
                VALUES (?, ?, ?)
            ''', (filename, report_hash, module_name))
            
            self.db.commit()
            return True
        except Exception as e:
            print(f"Error storing hash: {e}")
            return False
    
    def verify_report(self, filename):
        try:
            with open(filename, 'rb') as f:
                content = f.read()
            
            content_str = content.decode('latin-1', errors='ignore')
            
            hash_match = re.search(r'Hash: ([a-fA-F0-9]{64})', content_str)
            
            if not hash_match:
                return False, "No hash found in PDF"
            
            file_hash = hash_match.group(1)
            
            cursor = self.db.cursor()
            cursor.execute('''
                SELECT hash FROM report_hashes 
                WHERE filename=? OR hash LIKE ?
                ORDER BY created_at DESC LIMIT 1
            ''', (filename, f"{file_hash[:20]}%"))
            
            result = cursor.fetchone()
            
            if result:
                db_hash = result['hash']
                if file_hash == db_hash:
                    return True, "✓ Report is authentic (not tampered)"
                else:
                    return False, "✗ Report has been modified!"
            else:
                return False, "✗ Report not found in database"
                
        except Exception as e:
            return False, f"Verification error: {str(e)}"

class SimpleBlockchainVerifier:
    def __init__(self, db_conn):
        self.db = db_conn
        self.init_blockchain_table()
    
    def init_blockchain_table(self):
        cursor = self.db.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS blockchain_ledger (
                block_id INTEGER PRIMARY KEY AUTOINCREMENT,
                previous_hash TEXT,
                current_hash TEXT NOT NULL,
                data_hash TEXT NOT NULL,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                module_name TEXT,
                action_type TEXT,
                description TEXT
            )
        ''')
        self.db.commit()
    
    def add_to_blockchain(self, data, module_name="", action_type="", description=""):
        try:
            data_hash = hashlib.sha256(str(data).encode()).hexdigest()
            
            cursor = self.db.cursor()
            cursor.execute("SELECT current_hash FROM blockchain_ledger ORDER BY block_id DESC LIMIT 1")
            result = cursor.fetchone()
            previous_hash = result['current_hash'] if result else "0" * 64
            
            combined = previous_hash + data_hash
            current_hash = hashlib.sha256(combined.encode()).hexdigest()
            
            cursor.execute('''
                INSERT INTO blockchain_ledger 
                (previous_hash, current_hash, data_hash, module_name, action_type, description)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (previous_hash, current_hash, data_hash, module_name, action_type, description))
            
            self.db.commit()
            
            return {
                'block_id': cursor.lastrowid,
                'current_hash': current_hash,
                'data_hash': data_hash,
                'timestamp': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
        except Exception as e:
            print(f"Blockchain error: {e}")
            return None
    
    def verify_chain(self):
        try:
            cursor = self.db.cursor()
            cursor.execute("SELECT * FROM blockchain_ledger ORDER BY block_id")
            blocks = cursor.fetchall()
            
            if not blocks:
                return True, "Chain is empty"
            
            previous_hash = "0" * 64
            
            for block in blocks:
                combined = previous_hash + block['data_hash']
                calculated_hash = hashlib.sha256(combined.encode()).hexdigest()
                
                if calculated_hash != block['current_hash']:
                    return False, f"Chain broken at block {block['block_id']}"
                
                previous_hash = block['current_hash']
            
            return True, f"✓ Blockchain verified ({len(blocks)} blocks intact)"
            
        except Exception as e:
            return False, f"Verification error: {str(e)}"
    
    def get_latest_hash(self):
        cursor = self.db.cursor()
        cursor.execute("SELECT current_hash FROM blockchain_ledger ORDER BY block_id DESC LIMIT 1")
        result = cursor.fetchone()
        return result['current_hash'] if result else None

class HardeningApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Enterprise Linux Hardening Tool v2.0")
        self.root.geometry("1400x850")
        self.root.configure(bg=BG_COLOR)
        self.current_module_index = 0
        self.running_proc = False
        self.current_action = None
        self.filter_text = ""
        self.selected_policies_mode = False  # Track if we're in selected policies mode
        self.selected_policies = {}  # Store selected policies by module

        try:
            os.makedirs(os.path.dirname(DB_FILE), exist_ok=True)
            self.conn = sqlite3.connect(DB_FILE, check_same_thread=False)
            self.conn.row_factory = sqlite3.Row
            self.init_database()
        except Exception as e:
            messagebox.showerror("Database Error", f"Cannot connect to database:\n{DB_FILE}\nError: {e}")
            self.conn = None

        self.left_frame = tk.Frame(root, bg=SIDEBAR_COLOR, width=260, relief=tk.RAISED, borderwidth=2)
        self.left_frame.pack(side=tk.LEFT, fill=tk.Y)
        self.left_frame.pack_propagate(False)
        
        header_left = tk.Frame(self.left_frame, bg=HEADER_COLOR, height=60)
        header_left.pack(fill=tk.X)
        tk.Label(header_left, text="HARDENING", fg="white", bg=HEADER_COLOR, 
                font=("Segoe UI", 16, "bold")).pack(side=tk.LEFT, padx=15, pady=10)
        
        modules_frame = LabelFrame(self.left_frame, text=" MODULES ", font=("Segoe UI", 11, "bold"),
                                  bg=SIDEBAR_COLOR, fg=TEXT_COLOR, relief=tk.FLAT, borderwidth=0)
        modules_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
        
        self.module_buttons = []
        for idx, (title, _) in enumerate(MODULES):
            btn_frame = tk.Frame(modules_frame, bg=SIDEBAR_COLOR)
            btn_frame.pack(fill=tk.X, padx=5, pady=2)
            
            b = tk.Button(btn_frame, text=title, width=22, height=1,
                         font=("Segoe UI", 9),
                         bg=SIDEBAR_COLOR, fg=TEXT_COLOR,
                         relief=tk.FLAT, anchor="w",
                         command=lambda i=idx: self.select_module(i))
            b.pack(side=tk.LEFT, padx=5)
            
            b.bind("<Enter>", lambda e, b=b: b.configure(bg="#90caf9"))
            b.bind("<Leave>", lambda e, b=b: b.configure(bg=SIDEBAR_COLOR))
            
            self.module_buttons.append(b)

        # Add policy selector section
        selector_frame = LabelFrame(self.left_frame, text=" POLICY SELECTOR ", font=("Segoe UI", 11, "bold"),
                                   bg=SIDEBAR_COLOR, fg=TEXT_COLOR, relief=tk.FLAT, borderwidth=0)
        selector_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
        
        # Policy selector button
        self.policy_selector_btn = tk.Button(selector_frame, text="🔧 Select Policies", width=20, height=1,
                                           font=("Segoe UI", 9, "bold"),
                                           bg="#9c27b0", fg="white",
                                           relief=tk.RAISED, borderwidth=1,
                                           command=self.open_policy_selector)
        self.policy_selector_btn.pack(padx=10, pady=3)
        self.policy_selector_btn.bind("<Enter>", lambda e, b=self.policy_selector_btn: b.configure(relief=tk.SUNKEN))
        self.policy_selector_btn.bind("<Leave>", lambda e, b=self.policy_selector_btn: b.configure(relief=tk.RAISED))
        
        # Normal mode button (initially hidden)
        self.normal_mode_btn = tk.Button(selector_frame, text="📋 Show All Policies", width=20, height=1,
                                        font=("Segoe UI", 9),
                                        bg="#4caf50", fg="white",
                                        relief=tk.RAISED, borderwidth=1,
                                        command=self.show_all_policies)
        self.normal_mode_btn.pack(padx=10, pady=3)
        self.normal_mode_btn.pack_forget()
        self.normal_mode_btn.bind("<Enter>", lambda e, b=self.normal_mode_btn: b.configure(relief=tk.SUNKEN))
        self.normal_mode_btn.bind("<Leave>", lambda e, b=self.normal_mode_btn: b.configure(relief=tk.RAISED))

        actions_frame = LabelFrame(self.left_frame, text=" ACTIONS ", font=("Segoe UI", 11, "bold"),
                                  bg=SIDEBAR_COLOR, fg=TEXT_COLOR, relief=tk.FLAT, borderwidth=0)
        actions_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
        
        action_buttons = [
            ("Scan All", "#4caf50", lambda: self.start_action_all("scan")),
            ("Fix All", "#2196f3", lambda: self.start_action_all("fix")),
            ("Rollback All", "#f44336", self.rollback_all_modules),
        ]
        
        for text, color, command in action_buttons:
            btn = tk.Button(actions_frame, text=text, width=20, height=1,
                          font=("Segoe UI", 9, "bold"),
                          bg=color, fg="white",
                          relief=tk.RAISED, borderwidth=1,
                          command=command)
            btn.pack(padx=10, pady=3)
            btn.bind("<Enter>", lambda e, b=btn: b.configure(relief=tk.SUNKEN))
            btn.bind("<Leave>", lambda e, b=btn: b.configure(relief=tk.RAISED))

        # Add Clean DB button
        cleanup_btn = tk.Button(actions_frame, text="Clean DB", width=20, height=1,
                              font=("Segoe UI", 9),
                              bg="#ff9800", fg="white",
                              relief=tk.RAISED, borderwidth=1,
                              command=self.cleanup_duplicates)
        cleanup_btn.pack(padx=10, pady=3)
        cleanup_btn.bind("<Enter>", lambda e, b=cleanup_btn: b.configure(relief=tk.SUNKEN))
        cleanup_btn.bind("<Leave>", lambda e, b=cleanup_btn: b.configure(relief=tk.RAISED))

        stats_frame = LabelFrame(self.left_frame, text=" SYSTEM INFO ", font=("Segoe UI", 11, "bold"),
                                bg=SIDEBAR_COLOR, fg=TEXT_COLOR, relief=tk.FLAT, borderwidth=0)
        stats_frame.pack(fill=tk.X, padx=10, pady=(20, 10))
        
        self.stats_vars = {
            "Total Modules": tk.StringVar(value="9"),
            "Total Policies": tk.StringVar(value="0"),
            "Last Scan": tk.StringVar(value="Never"),
            "Database": tk.StringVar(value="Connected" if self.conn else "Error"),
        }
        
        for label, var in self.stats_vars.items():
            stat_frame = tk.Frame(stats_frame, bg=SIDEBAR_COLOR)
            stat_frame.pack(fill=tk.X, padx=5, pady=2)
            tk.Label(stat_frame, text=label+":", bg=SIDEBAR_COLOR, fg=TEXT_COLOR,
                    font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=5)
            tk.Label(stat_frame, textvariable=var, bg=SIDEBAR_COLOR, fg="#1e88e5",
                    font=("Segoe UI", 9, "bold")).pack(side=tk.RIGHT, padx=5)

        self.main_frame = tk.Frame(root, bg=BG_COLOR)
        self.main_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        header_frame = tk.Frame(self.main_frame, bg=HEADER_COLOR, height=70)
        header_frame.pack(fill=tk.X)
        
        title_frame = tk.Frame(header_frame, bg=HEADER_COLOR)
        title_frame.pack(side=tk.LEFT, padx=20, pady=10)
        
        tk.Label(title_frame, text="Enterprise Linux", fg="white", bg=HEADER_COLOR,
                font=("Segoe UI", 12)).pack(anchor="w")
        tk.Label(title_frame, text="HARDENING TOOL", fg="white", bg=HEADER_COLOR,
                font=("Segoe UI", 18, "bold")).pack(anchor="w")
        
        self.current_module_var = tk.StringVar(value="Select a Module")
        current_module_label = tk.Label(header_frame, textvariable=self.current_module_var,
                                       fg="white", bg=HEADER_COLOR, font=("Segoe UI", 11, "bold"))
        current_module_label.pack(side=tk.RIGHT, padx=20, pady=10)

        control_frame = tk.Frame(self.main_frame, bg=BG_COLOR)
        control_frame.pack(fill=tk.X, padx=20, pady=(15, 10))
        
        btn_frame = tk.Frame(control_frame, bg=BG_COLOR)
        btn_frame.pack(side=tk.LEFT)
        
        self.scan_btn = self.create_action_button(btn_frame, "Scan", "#4caf50", lambda: self.start_action("scan"))
        self.scan_btn.pack(side=tk.LEFT, padx=3)
        
        self.fix_btn = self.create_action_button(btn_frame, "Fix", "#2196f3", lambda: self.start_action("fix"))
        self.fix_btn.pack(side=tk.LEFT, padx=3)
        
        self.rollback_btn = self.create_action_button(btn_frame, "Rollback", "#f44336", self.start_module_rollback)
        self.rollback_btn.pack(side=tk.LEFT, padx=3)
        
        self.pdf_btn = self.create_action_button(btn_frame, "📄 PDF Report", "#673ab7", self.generate_pdf_report)
        self.pdf_btn.pack(side=tk.LEFT, padx=3)
        
        self.blockchain_btn = self.create_action_button(btn_frame, "🔗 Verify", "#009688", self.verify_blockchain)
        self.blockchain_btn.pack(side=tk.LEFT, padx=3)
        
        search_frame = tk.Frame(control_frame, bg=BG_COLOR)
        search_frame.pack(side=tk.RIGHT)
        
        tk.Label(search_frame, text="Search:", bg=BG_COLOR, fg=TEXT_COLOR,
                font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 5))
        
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(search_frame, textvariable=self.search_var, width=25,
                                    font=("Segoe UI", 9), relief=tk.SUNKEN, borderwidth=1)
        self.search_entry.pack(side=tk.LEFT, padx=5)
        self.search_entry.bind("<KeyRelease>", self.filter_table)
        
        search_btn = tk.Button(search_frame, text="🔍", width=2, height=1,
                              bg=BUTTON_COLOR, fg="white", font=("Segoe UI", 9),
                              command=self.filter_table)
        search_btn.pack(side=tk.LEFT, padx=2)
        
        clear_btn = tk.Button(search_frame, text="Clear", width=6, height=1,
                             bg="#9e9e9e", fg="white", font=("Segoe UI", 9),
                             command=self.clear_filter)
        clear_btn.pack(side=tk.LEFT, padx=2)

        self.progress_frame = tk.Frame(self.main_frame, bg=BG_COLOR)
        self.progress_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        
        self.progress_label = tk.Label(self.progress_frame, text="", bg=BG_COLOR, fg=TEXT_COLOR,
                                      font=("Segoe UI", 9))
        self.progress_label.pack(anchor="w")
        
        self.progress_bar = ttk.Progressbar(self.progress_frame, mode='indeterminate', length=400)
        self.progress_bar.pack(fill=tk.X, pady=(2, 0))
        
        self.progress_frame.pack_forget()

        self.tabs = ttk.Notebook(self.main_frame)
        self.tabs.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        style = ttk.Style()
        style.configure("TNotebook", background=BG_COLOR)
        style.configure("TNotebook.Tab", background=SIDEBAR_COLOR, foreground=TEXT_COLOR)
        style.map("TNotebook.Tab", background=[("selected", HEADER_COLOR)])

        self.scan_frame = tk.Frame(self.tabs, bg=BG_COLOR)
        self.tabs.add(self.scan_frame, text=" Scan Results ")
        
        table_controls = tk.Frame(self.scan_frame, bg=BG_COLOR)
        table_controls.pack(fill=tk.X, padx=5, pady=(5, 0))
        
        self.view_status_label = tk.Label(table_controls, text="Showing: All policies", bg=BG_COLOR, fg=TEXT_COLOR,
                                         font=("Segoe UI", 10, "bold"))
        self.view_status_label.pack(side=tk.LEFT)
        
        export_btn = tk.Button(table_controls, text="Export to Excel", width=12, height=1,
                              bg="#4caf50", fg="white", font=("Segoe UI", 9),
                              command=self.export_to_excel)
        export_btn.pack(side=tk.RIGHT, padx=5)
        
        refresh_btn = tk.Button(table_controls, text="Refresh", width=8, height=1,
                               bg=BUTTON_COLOR, fg="white", font=("Segoe UI", 9),
                               command=self.refresh_table)
        refresh_btn.pack(side=tk.RIGHT, padx=5)
        
        tree_frame = tk.Frame(self.scan_frame, bg=BG_COLOR)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        y_scrollbar = ttk.Scrollbar(tree_frame)
        y_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        x_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        x_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        columns = ("Policy ID", "Policy Name", "Expected", "Current", "Status")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings",
                                yscrollcommand=y_scrollbar.set,
                                xscrollcommand=x_scrollbar.set)
        
        col_widths = [120, 250, 150, 150, 100]
        for col, width in zip(columns, col_widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, minwidth=80, anchor="center")
        
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        y_scrollbar.config(command=self.tree.yview)
        x_scrollbar.config(command=self.tree.xview)
        
        self.tree.bind("<Double-1>", self.show_policy_details)

        self.console_frame = tk.Frame(self.tabs, bg=BG_COLOR)
        self.tabs.add(self.console_frame, text=" Console Output ")
        
        console_controls = tk.Frame(self.console_frame, bg=BG_COLOR)
        console_controls.pack(fill=tk.X, padx=5, pady=(5, 0))
        
        tk.Label(console_controls, text="Real-time Output", bg=BG_COLOR, fg=TEXT_COLOR,
                font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        
        clear_console_btn = tk.Button(console_controls, text="Clear Console", width=12, height=1,
                                     bg="#9e9e9e", fg="white", font=("Segoe UI", 9),
                                     command=self.clear_console)
        clear_console_btn.pack(side=tk.RIGHT, padx=5)
        
        save_log_btn = tk.Button(console_controls, text="Save Log", width=10, height=1,
                                bg=BUTTON_COLOR, fg="white", font=("Segoe UI", 9),
                                command=self.save_log)
        save_log_btn.pack(side=tk.RIGHT, padx=5)
        
        self.output_box = scrolledtext.ScrolledText(self.console_frame, wrap=tk.WORD,
                                                   font=("Consolas", 10), bg="#fafafa",
                                                   relief=tk.SUNKEN, borderwidth=1)
        self.output_box.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.output_box.tag_config("pass", foreground=COLOR_PASS)
        self.output_box.tag_config("fail", foreground=COLOR_FAIL)
        self.output_box.tag_config("warn", foreground=COLOR_WARN)
        self.output_box.tag_config("info", foreground=COLOR_INFO)
        self.output_box.tag_config("fixed", foreground=COLOR_FIXED)
        self.output_box.tag_config("manual", foreground=COLOR_MANUAL)
        self.output_box.tag_config("running", foreground=COLOR_RUNNING)
        self.output_box.tag_config("normal", foreground=COLOR_NORMAL)

        self.history_frame = tk.Frame(self.tabs, bg=BG_COLOR)
        self.tabs.add(self.history_frame, text=" Fix History ")
        
        history_controls = tk.Frame(self.history_frame, bg=BG_COLOR)
        history_controls.pack(fill=tk.X, padx=5, pady=(5, 0))
        
        tk.Label(history_controls, text="Fix History", bg=BG_COLOR, fg=TEXT_COLOR,
                font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        
        refresh_history_btn = tk.Button(history_controls, text="Refresh", width=8, height=1,
                                       bg=BUTTON_COLOR, fg="white", font=("Segoe UI", 9),
                                       command=self.load_fix_history)
        refresh_history_btn.pack(side=tk.RIGHT, padx=5)
        
        history_tree_frame = tk.Frame(self.history_frame, bg=BG_COLOR)
        history_tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        history_y_scrollbar = ttk.Scrollbar(history_tree_frame)
        history_y_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        history_x_scrollbar = ttk.Scrollbar(history_tree_frame, orient=tk.HORIZONTAL)
        history_x_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        history_columns = ("Policy ID", "Policy Name", "Original", "Current", "Status", "Timestamp")
        self.history_tree = ttk.Treeview(history_tree_frame, columns=history_columns, show="headings",
                                        yscrollcommand=history_y_scrollbar.set,
                                        xscrollcommand=history_x_scrollbar.set)
        
        history_col_widths = [100, 200, 120, 120, 80, 150]
        for col, width in zip(history_columns, history_col_widths):
            self.history_tree.heading(col, text=col)
            self.history_tree.column(col, width=width, minwidth=80, anchor="center")
        
        self.history_tree.pack(fill=tk.BOTH, expand=True)
        
        history_y_scrollbar.config(command=self.history_tree.yview)
        history_x_scrollbar.config(command=self.history_tree.xview)

        summary_frame = tk.Frame(self.main_frame, bg="#e3f2fd", bd=1, relief=tk.RAISED, borderwidth=1)
        summary_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        
        left_sum = tk.Frame(summary_frame, bg="#e3f2fd")
        left_sum.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=15, pady=10)
        
        tk.Label(left_sum, text="COMPLIANCE SUMMARY", font=("Segoe UI", 11, "bold"),
                bg="#e3f2fd", fg=TEXT_COLOR).pack(anchor="w", pady=(0, 5))
        
        self.total_checks_var = tk.StringVar(value="Total Checks: 0")
        tk.Label(left_sum, textvariable=self.total_checks_var, bg="#e3f2fd", fg=TEXT_COLOR,
                font=("Segoe UI", 9)).pack(anchor="w")
        
        self.compliance_var = tk.StringVar(value="Compliance: 0%")
        tk.Label(left_sum, textvariable=self.compliance_var, bg="#e3f2fd", fg=TEXT_COLOR,
                font=("Segoe UI", 9)).pack(anchor="w")
        
        self.pass_var = tk.StringVar(value="Passed: 0")
        tk.Label(left_sum, textvariable=self.pass_var, bg="#e3f2fd", fg=COLOR_PASS,
                font=("Segoe UI", 9)).pack(anchor="w")
        
        self.fail_var = tk.StringVar(value="Failed: 0")
        tk.Label(left_sum, textvariable=self.fail_var, bg="#e3f2fd", fg=COLOR_FAIL,
                font=("Segoe UI", 9)).pack(anchor="w")
        
        right_sum = tk.Frame(summary_frame, bg="#e3f2fd")
        right_sum.pack(side=tk.RIGHT, padx=15, pady=10)
        
        self.hash_var = tk.StringVar(value="Report ID: Not generated")
        tk.Label(right_sum, textvariable=self.hash_var, bg="#e3f2fd", fg="#1e88e5",
                font=("Segoe UI", 8)).pack(anchor="e", pady=1)
        
        action_btn_frame = tk.Frame(right_sum, bg="#e3f2fd")
        action_btn_frame.pack(anchor="e", pady=(5, 0))
        
        tk.Button(action_btn_frame, text="Export All", width=10, height=1,
                 bg="#4caf50", fg="white", font=("Segoe UI", 9),
                 command=self.export_all_data).pack(side=tk.LEFT, padx=2)
        
        tk.Button(action_btn_frame, text="Dashboard", width=10, height=1,
                 bg=BUTTON_COLOR, fg="white", font=("Segoe UI", 9),
                 command=self.show_dashboard).pack(side=tk.LEFT, padx=2)
        
        tk.Button(action_btn_frame, text="About", width=10, height=1,
                 bg="#9e9e9e", fg="white", font=("Segoe UI", 9),
                 command=self.show_about).pack(side=tk.LEFT, padx=2)

        self.reset_counters()
        if self.conn:
            self.select_module(0)
        else:
            self.append_console("[ERROR] Database not connected. Check DB_FILE path.")

    def init_database(self):
        cursor = self.conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS scan_results (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    module_name TEXT NOT NULL,
    policy_id TEXT NOT NULL,
    policy_name TEXT NOT NULL,
    expected_value TEXT NOT NULL,
    current_value TEXT NOT NULL,
    status TEXT NOT NULL,
    scan_timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(module_name, policy_id)
)
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS selected_policies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                module_name TEXT NOT NULL,
                policy_id TEXT NOT NULL,
                selected_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(module_name, policy_id)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS fix_history (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    module_name TEXT NOT NULL,
    policy_id TEXT NOT NULL,
    policy_name TEXT NOT NULL,
    expected_value TEXT NOT NULL,
    original_value TEXT NOT NULL,
    current_value TEXT NOT NULL,
    status TEXT NOT NULL,
    fix_timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
    rollback_executed TEXT DEFAULT 'NO',
    UNIQUE(module_name, policy_id)
)
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS blockchain_ledger (
                block_id INTEGER PRIMARY KEY AUTOINCREMENT,
                previous_hash TEXT,
                current_hash TEXT NOT NULL,
                data_hash TEXT NOT NULL,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                module_name TEXT,
                action_type TEXT,
                description TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS report_hashes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT NOT NULL,
                hash TEXT NOT NULL,
                module_name TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        self.conn.commit()
        
        blockchain = SimpleBlockchainVerifier(self.conn)
        cursor.execute("SELECT COUNT(*) as count FROM blockchain_ledger")
        if cursor.fetchone()['count'] == 0:
            blockchain.add_to_blockchain(
                data="GENESIS_BLOCK",
                module_name="SYSTEM",
                action_type="INIT",
                description="Initial blockchain genesis block"
            )

    def create_action_button(self, parent, text, color, command):
        btn = tk.Button(parent, text=text, width=10, height=1,
                       font=("Segoe UI", 10, "bold"),
                       bg=color, fg="white",
                       relief=tk.RAISED, borderwidth=2,
                       command=command)
        btn.bind("<Enter>", lambda e, b=btn: b.configure(relief=tk.SUNKEN))
        btn.bind("<Leave>", lambda e, b=btn: b.configure(relief=tk.RAISED))
        return btn

    def select_module(self, idx):
        self.current_module_index = idx
        module_name = MODULES[idx][0]
        self.current_module_var.set(f"Module: {module_name}")
        
        for i, b in enumerate(self.module_buttons):
            if i == idx:
                b.configure(bg="#90caf9", relief=tk.SUNKEN)
            else:
                b.configure(bg=SIDEBAR_COLOR, relief=tk.FLAT)
        
        if self.conn:
            self.load_scan_results(module_name)
            self.load_fix_history()
        
        self.clear_console()
        self.reset_counters()
        self.clear_filter()
    
    def load_scan_results(self, module_name):
        """Load only the latest scan results for each policy"""
        self.tree.delete(*self.tree.get_children())
        
        if not self.conn:
            self.append_console("[ERROR] Database not connected")
            return
        
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='scan_results'")
            if not cursor.fetchone():
                self.append_console(f"[INFO] scan_results table doesn't exist yet. Run a scan first.")
                return
            
            # Check if we have selected policies for this module
            cursor.execute("SELECT policy_id FROM selected_policies WHERE module_name=?", (module_name,))
            selected_policies = [row['policy_id'] for row in cursor.fetchall()]
            
            if selected_policies and self.selected_policies_mode:
                # Build query with selected policies filter
                placeholders = ','.join(['?'] * len(selected_policies))
                query = f"""
                    SELECT s1.policy_id, s1.policy_name, s1.expected_value, s1.current_value, s1.status 
                    FROM scan_results s1
                    INNER JOIN (
                        SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                        FROM scan_results 
                        WHERE module_name=?
                        GROUP BY policy_id, module_name
                    ) s2 ON s1.policy_id = s2.policy_id 
                        AND s1.module_name = s2.module_name 
                        AND s1.scan_timestamp = s2.max_timestamp
                    WHERE s1.module_name=? AND s1.policy_id IN ({placeholders})
                    ORDER BY s1.policy_id
                """
                params = [module_name, module_name] + selected_policies
                self.view_status_label.config(text=f"Showing: {len(selected_policies)} selected policies")
            else:
                # Show all policies
                query = """
                    SELECT s1.policy_id, s1.policy_name, s1.expected_value, s1.current_value, s1.status 
                    FROM scan_results s1
                    INNER JOIN (
                        SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                        FROM scan_results 
                        WHERE module_name=?
                        GROUP BY policy_id, module_name
                    ) s2 ON s1.policy_id = s2.policy_id 
                        AND s1.module_name = s2.module_name 
                        AND s1.scan_timestamp = s2.max_timestamp
                    WHERE s1.module_name=?
                    ORDER BY s1.policy_id
                """
                params = [module_name, module_name]
                self.view_status_label.config(text="Showing: All policies")
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            if not rows:
                if selected_policies and self.selected_policies_mode:
                    self.append_console(f"[INFO] No scan results found for selected policies in {module_name}. Run a scan first.")
                else:
                    self.append_console(f"[INFO] No scan results found for {module_name}. Run a scan first.")
                return
            
            self.stats_vars["Total Policies"].set(str(len(rows)))
            
            converted_rows = []
            for row in rows:
                converted_row = (
                    row['policy_id'],
                    row['policy_name'],
                    row['expected_value'],
                    row['current_value'],
                    row['status']
                )
                converted_rows.append(converted_row)
            
            filtered_rows = converted_rows
            if self.filter_text:
                filtered_rows = [
                    r for r in converted_rows 
                    if any(self.filter_text.lower() in str(field).lower() 
                          for field in r)
                ]
            
            for r in filtered_rows:
                status = r[4].lower() if r[4] else "normal"
                self.tree.insert("", tk.END, values=r, tags=(status,))
            
            self.update_counts_from_db(converted_rows)
            self.update_summary()
            
            for status, color in [
                ("pass", COLOR_PASS), ("fail", COLOR_FAIL),
                ("manual", COLOR_MANUAL), ("fixed", COLOR_FIXED)
            ]:
                self.tree.tag_configure(status, foreground=color)
            
        except Exception as e:
            self.append_console(f"[ERROR] Failed to load scan results: {e}")
    
    def load_fix_history(self):
        self.history_tree.delete(*self.history_tree.get_children())
        
        if not self.conn or self.current_module_index >= len(MODULES):
            return
        
        module_name = MODULES[self.current_module_index][0]
        
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='fix_history'")
            if not cursor.fetchone():
                return
            
            cursor.execute("""
                SELECT policy_id, policy_name, original_value, current_value, status, fix_timestamp
                FROM fix_history 
                WHERE module_name=?
                ORDER BY fix_timestamp DESC
                LIMIT 100
            """, (module_name,))
            rows = cursor.fetchall()
            
            for r in rows:
                converted_row = (
                    r['policy_id'],
                    r['policy_name'],
                    r['original_value'],
                    r['current_value'],
                    r['status'],
                    r['fix_timestamp']
                )
                self.history_tree.insert("", tk.END, values=converted_row)
                
        except Exception as e:
            self.append_console(f"[ERROR] Failed to load fix history: {e}")

    def open_policy_selector(self):
        """Open dialog to select policies with checkboxes"""
        if not self.conn:
            messagebox.showerror("Error", "Database not connected")
            return
        
        module_name = MODULES[self.current_module_index][0] if self.current_module_index < len(MODULES) else "Unknown"
        
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Policy Selector - {module_name}")
        dialog.geometry("800x600")
        dialog.configure(bg=BG_COLOR)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Header
        header_frame = tk.Frame(dialog, bg=HEADER_COLOR)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Label(header_frame, text=f"Select Policies for {module_name}", 
                 fg="white", bg=HEADER_COLOR, font=("Segoe UI", 12, "bold")).pack(pady=10)
        
        # Instructions
        info_frame = tk.Frame(dialog, bg=BG_COLOR)
        info_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        tk.Label(info_frame, text="✓ Check policies you want to display\n✗ Uncheck to hide from display\n\nNote: Scripts still run all policies", 
                 bg=BG_COLOR, fg=TEXT_COLOR, font=("Segoe UI", 9), justify=tk.LEFT).pack(anchor="w")
        
        # Create main frame with scrollbar
        main_frame = tk.Frame(dialog, bg=BG_COLOR)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Create canvas and scrollbar
        canvas = tk.Canvas(main_frame, bg=BG_COLOR, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=BG_COLOR)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Control buttons frame
        control_frame = tk.Frame(scrollable_frame, bg=BG_COLOR)
        control_frame.pack(fill=tk.X, pady=(0, 10))
        
        select_all_btn = tk.Button(control_frame, text="Select All", width=12,
                                  bg="#4caf50", fg="white", font=("Segoe UI", 9),
                                  command=lambda: self.select_all_checkboxes(checkboxes, True))
        select_all_btn.pack(side=tk.LEFT, padx=5)
        
        select_none_btn = tk.Button(control_frame, text="Select None", width=12,
                                   bg="#f44336", fg="white", font=("Segoe UI", 9),
                                   command=lambda: self.select_all_checkboxes(checkboxes, False))
        select_none_btn.pack(side=tk.LEFT, padx=5)
        
        # Get current policies from database
        try:
            cursor = self.conn.cursor()
            cursor.execute("""
                SELECT DISTINCT s1.policy_id, s1.policy_name
                FROM scan_results s1
                INNER JOIN (
                    SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                    FROM scan_results 
                    WHERE module_name=?
                    GROUP BY policy_id, module_name
                ) s2 ON s1.policy_id = s2.policy_id 
                    AND s1.module_name = s2.module_name 
                    AND s1.scan_timestamp = s2.max_timestamp
                WHERE s1.module_name=?
                ORDER BY s1.policy_id
            """, (module_name, module_name))
            
            policies = cursor.fetchall()
            
            if not policies:
                tk.Label(scrollable_frame, text="No policies found. Run a scan first.", 
                        bg=BG_COLOR, fg=TEXT_COLOR, font=("Segoe UI", 10)).pack(pady=20)
            else:
                # Get currently selected policies
                cursor.execute("SELECT policy_id FROM selected_policies WHERE module_name=?", (module_name,))
                selected_ids = {row['policy_id'] for row in cursor.fetchall()}
                
                # Create checkboxes
                checkboxes = {}
                checkbox_frame = tk.Frame(scrollable_frame, bg=BG_COLOR)
                checkbox_frame.pack(fill=tk.BOTH, expand=True)
                
                for policy in policies:
                    policy_id = policy['policy_id']
                    policy_name = policy['policy_name']
                    
                    var = tk.BooleanVar(value=(policy_id in selected_ids))
                    
                    cb_frame = tk.Frame(checkbox_frame, bg=BG_COLOR)
                    cb_frame.pack(fill=tk.X, padx=5, pady=2)
                    
                    cb = tk.Checkbutton(cb_frame, text=f"{policy_id}: {policy_name}", 
                                       variable=var, bg=BG_COLOR, fg=TEXT_COLOR,
                                       font=("Segoe UI", 9), anchor="w")
                    cb.pack(side=tk.LEFT, fill=tk.X, expand=True)
                    
                    checkboxes[policy_id] = var
                
                tk.Label(control_frame, text=f"Total: {len(policies)} policies", 
                        bg=BG_COLOR, fg=TEXT_COLOR, font=("Segoe UI", 9)).pack(side=tk.RIGHT, padx=10)
        
        except Exception as e:
            tk.Label(scrollable_frame, text=f"Error loading policies: {str(e)}", 
                    bg=BG_COLOR, fg=COLOR_FAIL, font=("Segoe UI", 10)).pack(pady=20)
            checkboxes = {}
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Button frame at bottom
        button_frame = tk.Frame(dialog, bg=BG_COLOR)
        button_frame.pack(fill=tk.X, padx=20, pady=10)
        
        def save_selections():
            try:
                cursor = self.conn.cursor()
                
                # Clear existing selections for this module
                cursor.execute("DELETE FROM selected_policies WHERE module_name=?", (module_name,))
                
                # Save new selections
                selected_count = 0
                for policy_id, var in checkboxes.items():
                    if var.get():  # If checked
                        cursor.execute("""
                            INSERT INTO selected_policies (module_name, policy_id) 
                            VALUES (?, ?)
                        """, (module_name, policy_id))
                        selected_count += 1
                
                self.conn.commit()
                dialog.destroy()
                
                # Enable selected policies mode
                self.selected_policies_mode = True
                self.show_normal_mode_button()
                
                # Refresh current view
                self.load_scan_results(module_name)
                
                if selected_count > 0:
                    messagebox.showinfo("Success", f"{selected_count} policies selected for display!")
                else:
                    messagebox.showinfo("Info", "No policies selected. Showing all policies.")
                    self.selected_policies_mode = False
                    self.hide_normal_mode_button()
                    self.load_scan_results(module_name)
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save selections: {str(e)}")
        
        save_btn = tk.Button(button_frame, text="Apply Selection", width=15,
                            bg="#2196f3", fg="white", font=("Segoe UI", 10, "bold"),
                            command=save_selections)
        save_btn.pack(side=tk.LEFT, padx=5)
        
        cancel_btn = tk.Button(button_frame, text="Cancel", width=15,
                              bg="#9e9e9e", fg="white", font=("Segoe UI", 10),
                              command=dialog.destroy)
        cancel_btn.pack(side=tk.RIGHT, padx=5)

    def select_all_checkboxes(self, checkboxes, select=True):
        """Select or deselect all checkboxes"""
        for var in checkboxes.values():
            var.set(select)

    def show_all_policies(self):
        """Return to showing all policies (normal mode)"""
        if not self.selected_policies_mode:
            return
        
        module_name = MODULES[self.current_module_index][0] if self.current_module_index < len(MODULES) else "Unknown"
        
        response = messagebox.askyesno("Show All Policies", 
            f"Return to showing ALL policies for {module_name}?")
        
        if response:
            self.selected_policies_mode = False
            self.hide_normal_mode_button()
            self.load_scan_results(module_name)

    def show_normal_mode_button(self):
        """Show the normal mode button"""
        self.normal_mode_btn.pack(padx=10, pady=3)
        self.policy_selector_btn.config(text="🔧 Change Selection")

    def hide_normal_mode_button(self):
        """Hide the normal mode button"""
        self.normal_mode_btn.pack_forget()
        self.policy_selector_btn.config(text="🔧 Select Policies")

    def filter_table(self, event=None):
        self.filter_text = self.search_var.get().strip()
        if self.current_module_index < len(MODULES):
            module_name = MODULES[self.current_module_index][0]
            self.load_scan_results(module_name)

    def clear_filter(self):
        self.search_var.set("")
        self.filter_text = ""
        if self.current_module_index < len(MODULES):
            module_name = MODULES[self.current_module_index][0]
            self.load_scan_results(module_name)

    def refresh_table(self):
        if self.current_module_index < len(MODULES):
            module_name = MODULES[self.current_module_index][0]
            self.load_scan_results(module_name)
            self.load_fix_history()

    def show_policy_details(self, event):
        selection = self.tree.selection()
        if not selection:
            return
        
        item = self.tree.item(selection[0])
        values = item['values']
        
        if not values or len(values) < 5:
            return
        
        details_window = tk.Toplevel(self.root)
        details_window.title("Policy Details")
        details_window.geometry("500x400")
        details_window.configure(bg=BG_COLOR)
        
        details_text = scrolledtext.ScrolledText(details_window, wrap=tk.WORD,
                                                font=("Consolas", 10), bg="#fafafa")
        details_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        details = f"""POLICY DETAILS
{"="*50}
Policy ID:     {values[0]}
Policy Name:   {values[1]}
Expected:      {values[2]}
Current:       {values[3]}
Status:        {values[4]}
{"="*50}
Timestamp:     {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Module:        {MODULES[self.current_module_index][0]}
        """
        
        details_text.insert(tk.END, details)
        details_text.config(state=tk.DISABLED)

    def append_console(self, line):
        line = line.rstrip("\n")
        tag = "normal"
        
        if "[PASS]" in line:
            tag = "pass"
        elif "[FAIL]" in line:
            tag = "fail"
        elif "[WARN]" in line:
            tag = "warn"
        elif "[INFO]" in line:
            tag = "info"
        elif "[FIXED]" in line:
            tag = "fixed"
        elif "[MANUAL]" in line:
            tag = "manual"
        elif any(x in line for x in ["Running", "Starting", "Executing"]):
            tag = "running"
        
        self.output_box.insert(tk.END, line + "\n", tag)
        self.output_box.see(tk.END)
        
        if tag == "pass":
            self.count_pass += 1
            self.total_checks += 1
        elif tag == "fail":
            self.count_fail += 1
            self.total_checks += 1
        elif tag == "manual":
            self.count_manual += 1
            self.total_checks += 1
        elif tag == "fixed":
            self.count_fixed += 1
        
        self.update_summary()

    def clear_console(self):
        self.output_box.delete(1.0, tk.END)
        self.reset_counters()
        self.update_summary()

    def save_log(self):
        txt = self.output_box.get(1.0, tk.END)
        if not txt.strip():
            return
        
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        module = MODULES[self.current_module_index][0] if self.current_module_index < len(MODULES) else "unknown"
        
        path = filedialog.asksaveasfilename(
            defaultextension=".log",
            filetypes=[("Log files", "*.log"), ("Text files", "*.txt")],
            initialfile=f"hardening_{module}_{timestamp}.log"
        )
        
        if not path:
            return
        
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write(f"Hardening Log - {module}\n")
                f.write(f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("="*80 + "\n\n")
                f.write(txt)
            
            messagebox.showinfo("Log Saved", f"Log saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Save Error", f"Failed to save log:\n{e}")

    def show_progress(self, message):
        self.progress_label.config(text=message)
        self.progress_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        self.progress_bar.start(10)
    
    def hide_progress(self):
        self.progress_bar.stop()
        self.progress_frame.pack_forget()
        self.progress_label.config(text="")
    
    def update_progress(self, value):
        if self.progress_bar["mode"] == "determinate":
            self.progress_bar["value"] = value
            self.root.update_idletasks()

    def generate_pdf_report(self):
        if not self.conn:
            messagebox.showerror("Error", "Database not connected")
            return
        
        choice = messagebox.askquestion("PDF Report", 
                                   "Generate report for current module only?\n\n"
                                   "Yes = Current module only\n"
                                   "No = All modules")
    
        module_name = None
        if choice == 'yes':
            if self.current_module_index < len(MODULES):
                module_name = MODULES[self.current_module_index][0]
            else:
                module_name = "Unknown"
    
        try:
            self.show_progress("Generating PDF report...")
            
            pdf_gen = PDFReportGenerator(self.conn)
            
            filename, report_hash = pdf_gen.generate_report(module_name)
            
            report_id = f"HARDEN-{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            blockchain = SimpleBlockchainVerifier(self.conn)
            tx = blockchain.add_to_blockchain(
                data=report_hash,
                module_name=module_name or "ALL",
                action_type="PDF_REPORT",
                description=f"PDF report generated: {os.path.basename(filename)} - ID: {report_id}"
            )
            
            self.hide_progress()
            
            if tx:
                msg = f"✅ PDF Report Generated Successfully!\n\n"
                msg += f"Report ID: {report_id}\n"
                msg += f"File: {filename}\n"
                msg += f"Blockchain TX: {tx['current_hash'][:16]}...\n"
                msg += f"Timestamp: {tx['timestamp']}"
            else:
                msg = f"✅ PDF Report Generated!\n\nReport ID: {report_id}\nFile: {filename}"
            
            self.hash_var.set(f"Report ID: {report_id}")
            
            if messagebox.askyesno("Report Generated", msg + "\n\nOpen PDF file?"):
                if os.name == 'nt':
                    os.startfile(filename)
                else:
                    os.system(f"xdg-open '{filename}'")
        
        except Exception as e:
            self.hide_progress()
            messagebox.showerror("PDF Generation Error", f"Failed to generate PDF:\n{str(e)}")

    def verify_blockchain(self):
        if not self.conn:
            messagebox.showerror("Error", "Database not connected")
            return
        
        options = ["Verify Blockchain Chain", "Verify PDF Report", "View Blockchain"]
        choice = tk.StringVar(value=options[0])
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Blockchain Verification")
        dialog.geometry("400x200")
        dialog.configure(bg=BG_COLOR)
        
        tk.Label(dialog, text="Select Verification Type:", 
                 bg=BG_COLOR, fg=TEXT_COLOR, font=("Segoe UI", 11)).pack(pady=10)
        
        for option in options:
            tk.Radiobutton(dialog, text=option, variable=choice, value=option,
                          bg=BG_COLOR, fg=TEXT_COLOR).pack(anchor="w", padx=30)
        
        def perform_verification():
            selected = choice.get()
            dialog.destroy()
            
            if selected == "Verify Blockchain Chain":
                self.verify_blockchain_chain()
            elif selected == "Verify PDF Report":
                self.verify_pdf_report()
            elif selected == "View Blockchain":
                self.view_blockchain()
        
        tk.Button(dialog, text="Verify", bg=BUTTON_COLOR, fg="white",
                 command=perform_verification).pack(pady=20)

    def verify_blockchain_chain(self):
        blockchain = SimpleBlockchainVerifier(self.conn)
        is_valid, message = blockchain.verify_chain()
        
        if is_valid:
            messagebox.showinfo("Blockchain Verified", f"✅ {message}")
        else:
            messagebox.showerror("Blockchain Error", f"❌ {message}")

    def verify_pdf_report(self):
        filename = filedialog.askopenfilename(
            title="Select PDF Report to Verify",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        pdf_gen = PDFReportGenerator(self.conn)
        is_valid, message = pdf_gen.verify_report(filename)
        
        if is_valid:
            messagebox.showinfo("Report Verified", f"✅ {message}")
        else:
            messagebox.showerror("Report Tampered", f"❌ {message}")

    def view_blockchain(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("""
                SELECT block_id, module_name, action_type, 
                       SUBSTR(current_hash, 1, 16) as short_hash, 
                       timestamp, description
                FROM blockchain_ledger 
                ORDER BY block_id DESC
                LIMIT 50
            """)
            rows = cursor.fetchall()
            
            window = tk.Toplevel(self.root)
            window.title("Blockchain Ledger")
            window.geometry("800x500")
            window.configure(bg=BG_COLOR)
            
            tk.Label(window, text="🔗 Blockchain Transaction Ledger", 
                    bg=BG_COLOR, fg=TEXT_COLOR, font=("Segoe UI", 14, "bold")).pack(pady=10)
            
            text = scrolledtext.ScrolledText(window, wrap=tk.WORD, 
                                            font=("Consolas", 9), bg="#fafafa")
            text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            if not rows:
                text.insert(tk.END, "No blockchain transactions yet.")
            else:
                text.insert(tk.END, f"Total Blocks: {len(rows)}\n")
                text.insert(tk.END, "="*80 + "\n\n")
                
                for row in rows:
                    text.insert(tk.END, f"Block #{row['block_id']}\n")
                    text.insert(tk.END, f"  Module: {row['module_name']}\n")
                    text.insert(tk.END, f"  Action: {row['action_type']}\n")
                    text.insert(tk.END, f"  Hash: {row['short_hash']}...\n")
                    text.insert(tk.END, f"  Time: {row['timestamp']}\n")
                    text.insert(tk.END, f"  Desc: {row['description']}\n")
                    text.insert(tk.END, "-"*40 + "\n")
            
            text.config(state=tk.DISABLED)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load blockchain: {str(e)}")

    def start_action(self, action):
        if self.running_proc:
            messagebox.showwarning("Running", "Another action is running. Please wait.")
            return
        
        idx = self.current_module_index
        if idx >= len(MODULES):
            return
        
        title, fname = MODULES[idx]
        
        if not os.path.exists(fname):
            self.append_console(f"[FAIL] Script not found: {fname}")
            return
        
        if action in ["fix", "rollback"]:
            ok = messagebox.askyesno("Confirm", f"Are you sure to {action} {title}?\nThis will modify system configuration.")
            if not ok:
                return
        
        self.current_action = action
        cmd = ["sudo", "bash", fname, action]
        
        if self.conn:
            try:
                blockchain = SimpleBlockchainVerifier(self.conn)
                blockchain.add_to_blockchain(
                    data=f"{title}_{action}",
                    module_name=title,
                    action_type=action.upper(),
                    description=f"Started {action} on {title} module"
                )
            except:
                pass
        
        self.append_console("="*80)
        self.append_console(f"{title} - {action.upper()} - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        self.show_progress(f"Running {action} on {title}...")
        self.running_proc = True
        self._set_buttons_state("disabled")
        
        t = threading.Thread(target=lambda: self._run_and_stream(cmd, title, action))
        t.daemon = True
        t.start()

    def start_action_all(self, action):
        if self.running_proc:
            messagebox.showwarning("Running", "Another action is running. Please wait.")
            return
        
        if action in ["fix", "rollback"]:
            ok = messagebox.askyesno("Confirm", 
                f"Run {action} on ALL {len(MODULES)} modules?\nThis will modify system configuration.")
            if not ok:
                return
        
        def run_next(idx=0):
            if idx < len(MODULES):
                self.current_module_index = idx
                self.clear_console()
                self.current_module_var.set(f"Module: {MODULES[idx][0]}")
                
                for i, b in enumerate(self.module_buttons):
                    if i == idx:
                        b.configure(bg="#90caf9", relief=tk.SUNKEN)
                    else:
                        b.configure(bg=SIDEBAR_COLOR, relief=tk.FLAT)
                
                title, fname = MODULES[idx]
                cmd = ["sudo", "bash", fname, action]
                
                self.append_console(f"Processing: {title} ({idx+1}/{len(MODULES)})")
                self.append_console("-"*60)
                
                self.show_progress(f"Processing {title} ({idx+1}/{len(MODULES)})...")
                
                def on_line(line):
                    self.root.after(0, lambda: self.append_console(line))
                
                def on_done(returncode):
                    self.append_console(f"[INFO] Finished {title} with exit code {returncode}")
                    self.append_console("="*80)
                    self.hide_progress()
                    
                    if self.conn:
                        self.root.after(100, lambda: self.load_scan_results(title))
                        self.root.after(100, lambda: self.load_fix_history())
                    
                    self.root.after(500, lambda: run_next(idx + 1))
                
                t = threading.Thread(target=lambda: run_command_stream(cmd, on_line, on_done))
                t.daemon = True
                t.start()
                
            else:
                self.append_console("[INFO] All modules completed!")
                self.hide_progress()
                self._set_buttons_state("normal")
                self.running_proc = False
                messagebox.showinfo("Complete", f"All modules {action} completed!")
                self.stats_vars["Last Scan"].set(datetime.datetime.now().strftime("%H:%M"))
        
        self.clear_console()
        self._set_buttons_state("disabled")
        self.running_proc = True
        run_next()

    def start_module_rollback(self):
        idx = self.current_module_index
        if idx >= len(MODULES):
            return
        
        title = MODULES[idx][0]
        rollback_script = ROLLBACK_SCRIPTS.get(title)
        
        if not rollback_script or not os.path.exists(rollback_script):
            self.append_console(f"[FAIL] Rollback script not found for {title}")
            messagebox.showerror("Error", f"Rollback script not found:\n{rollback_script or 'N/A'}")
            return
        
        ok = messagebox.askyesno("Confirm Rollback", 
            f"Rollback {title}?\nThis will undo all fixes made by this module.")
        if not ok:
            return
        
        cmd = ["sudo", "bash", rollback_script]
        
        self.append_console("="*80)
        self.append_console(f"ROLLBACK {title} - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        self.show_progress(f"Rolling back {title}...")
        self.running_proc = True
        self._set_buttons_state("disabled")
        
        t = threading.Thread(target=lambda: self._run_and_stream(cmd, f"{title} Rollback", "rollback"))
        t.daemon = True
        t.start()

    def rollback_all_modules(self):
        if self.running_proc:
            messagebox.showwarning("Running", "Another action is running. Please wait.")
            return
        
        ok = messagebox.askyesno("Confirm Full Rollback", 
            f"Rollback ALL {len(MODULES)} modules?\nThis will undo ALL fixes across the system!")
        if not ok:
            return
        
        def run_next(idx=0):
            if idx < len(MODULES):
                title = MODULES[idx][0]
                rollback_script = ROLLBACK_SCRIPTS.get(title)
                
                if rollback_script and os.path.exists(rollback_script):
                    self.append_console(f"Rolling back: {title} ({idx+1}/{len(MODULES)})")
                    self.show_progress(f"Rolling back {title} ({idx+1}/{len(MODULES)})...")
                    
                    cmd = ["sudo", "bash", rollback_script]
                    
                    def on_line(line):
                        self.root.after(0, lambda: self.append_console(line))
                    
                    def on_done(returncode):
                        self.append_console(f"[INFO] Finished rollback {title} with exit code {returncode}")
                        self.append_console("-"*60)
                        self.hide_progress()
                        self.root.after(100, lambda: run_next(idx + 1))
                    
                    t = threading.Thread(target=lambda: run_command_stream(cmd, on_line, on_done))
                    t.daemon = True
                    t.start()
                else:
                    self.append_console(f"[WARN] Rollback script not found for {title}")
                    self.root.after(100, lambda: run_next(idx + 1))
            
            else:
                self.append_console("[INFO] All rollbacks completed!")
                self.hide_progress()
                self._set_buttons_state("normal")
                self.running_proc = False
                messagebox.showinfo("Complete", "All rollbacks completed!")
        
        self.clear_console()
        self._set_buttons_state("disabled")
        self.running_proc = True
        run_next()

    def _run_and_stream(self, cmd, title, action):
        def on_line(line):
            self.root.after(0, lambda: self.append_console(line))
        
        def on_done(returncode):
            self.running_proc = False
            self.hide_progress()
            self._set_buttons_state("normal")
            
            if returncode == 0:
                status = "SUCCESS"
            else:
                status = "FAILED"
            
            self.append_console(f"[INFO] {title} finished with exit code {returncode} ({status})")
            
            if self.conn and self.current_module_index < len(MODULES):
                module_name = MODULES[self.current_module_index][0]
                self.root.after(500, lambda: self.load_scan_results(module_name))
                self.root.after(500, lambda: self.load_fix_history())
                
                try:
                    blockchain = SimpleBlockchainVerifier(self.conn)
                    blockchain.add_to_blockchain(
                        data=f"{module_name}_{action}_{returncode}",
                        module_name=module_name,
                        action_type=f"{action.upper()}_COMPLETE",
                        description=f"Completed {action} on {module_name} with code {returncode}"
                    )
                except:
                    pass
            
            if action == "scan":
                self.stats_vars["Last Scan"].set(datetime.datetime.now().strftime("%H:%M"))
        
        run_command_stream(cmd, on_line, on_done)

    def _set_buttons_state(self, state):
        for w in [self.scan_btn, self.fix_btn, self.rollback_btn, self.pdf_btn, self.blockchain_btn,
                  self.policy_selector_btn, self.normal_mode_btn]:
            try:
                w.configure(state=state)
            except:
                pass

    def export_to_excel(self):
        if not self.conn:
            messagebox.showerror("Database Error", "Database not connected.")
            return
        
        module_name = MODULES[self.current_module_index][0] if self.current_module_index < len(MODULES) else "Unknown"
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"hardening_{module_name}_{timestamp}.xlsx"
        )
        
        if not path:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Scan Results"
            
            headers = ["Policy ID", "Policy Name", "Expected Value", "Current Value", "Status", "Module", "Timestamp"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")
            
            cursor = self.conn.cursor()
            # Get only latest results for export
            cursor.execute("""
                SELECT s1.policy_id, s1.policy_name, s1.expected_value, s1.current_value, s1.status, s1.module_name, s1.scan_timestamp
                FROM scan_results s1
                INNER JOIN (
                    SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                    FROM scan_results 
                    WHERE module_name=?
                    GROUP BY policy_id, module_name
                ) s2 ON s1.policy_id = s2.policy_id 
                    AND s1.module_name = s2.module_name 
                    AND s1.scan_timestamp = s2.max_timestamp
                WHERE s1.module_name=?
                ORDER BY s1.policy_id
            """, (module_name, module_name))
            
            rows = cursor.fetchall()
            
            for row_num, row in enumerate(rows, 2):
                row_data = (
                    row['policy_id'],
                    row['policy_name'],
                    row['expected_value'],
                    row['current_value'],
                    row['status'],
                    row['module_name'],
                    row['scan_timestamp']
                )
                
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
                    
                    if col_num == 5:
                        if row_data[4] == "PASS":
                            ws.cell(row=row_num, column=col_num).fill = PatternFill(
                                start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
                        elif row_data[4] == "FAIL":
                            ws.cell(row=row_num, column=col_num).fill = PatternFill(
                                start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
                        elif row_data[4] == "MANUAL":
                            ws.cell(row=row_num, column=col_num).fill = PatternFill(
                                start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(path)
            
            messagebox.showinfo("Export Successful", 
                f"Exported {len(rows)} records to:\n{path}\n\nModule: {module_name}")
            
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export to Excel:\n{e}")

    def export_all_data(self):
        if not self.conn:
            messagebox.showerror("Database Error", "Database not connected.")
            return
        
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"hardening_complete_{timestamp}.xlsx"
        )
        
        if not path:
            return
        
        try:
            wb = Workbook()
            
            for module_name, _ in MODULES:
                if not wb.sheetnames:
                    ws = wb.active
                    ws.title = module_name[:31]
                else:
                    ws = wb.create_sheet(title=module_name[:31])
                
                headers = ["Policy ID", "Policy Name", "Expected", "Current", "Status", "Timestamp"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")
                
                cursor = self.conn.cursor()
                # Get only latest results for each module
                cursor.execute("""
                    SELECT s1.policy_id, s1.policy_name, s1.expected_value, s1.current_value, s1.status, s1.scan_timestamp
                    FROM scan_results s1
                    INNER JOIN (
                        SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                        FROM scan_results 
                        WHERE module_name=?
                        GROUP BY policy_id, module_name
                    ) s2 ON s1.policy_id = s2.policy_id 
                        AND s1.module_name = s2.module_name 
                        AND s1.scan_timestamp = s2.max_timestamp
                    WHERE s1.module_name=?
                    ORDER BY s1.policy_id
                """, (module_name, module_name))
                
                rows = cursor.fetchall()
                
                for row_num, row in enumerate(rows, 2):
                    row_data = (
                        row['policy_id'],
                        row['policy_name'],
                        row['expected_value'],
                        row['current_value'],
                        row['status'],
                        row['scan_timestamp']
                    )
                    
                    for col_num, value in enumerate(row_data, 1):
                        ws.cell(row=row_num, column=col_num, value=value)
            
            if len(wb.sheetnames) > 1 and wb["Sheet"].max_row == 1:
                std = wb["Sheet"]
                wb.remove(std)
            
            wb.save(path)
            messagebox.showinfo("Export Successful", f"All data exported to:\n{path}")
            
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export all data:\n{e}")

    def update_counts_from_db(self, rows):
        self.count_pass = sum(1 for r in rows if r[4] == "PASS")
        self.count_fail = sum(1 for r in rows if r[4] == "FAIL")
        self.count_manual = sum(1 for r in rows if r[4] == "MANUAL")
        self.count_fixed = sum(1 for r in rows if r[4] == "FIXED")
        self.total_checks = len(rows)

    def reset_counters(self):
        self.count_pass = 0
        self.count_fail = 0
        self.count_warn = 0
        self.count_info = 0
        self.count_fixed = 0
        self.count_manual = 0
        self.total_checks = 0

    def update_summary(self):
        self.total_checks_var.set(f"Total Checks: {self.total_checks}")
        self.pass_var.set(f"Passed: {self.count_pass}")
        self.fail_var.set(f"Failed: {self.count_fail}")
        
        total_for_compliance = self.count_pass + self.count_fail
        if total_for_compliance > 0:
            pct = int((self.count_pass / total_for_compliance) * 100)
            self.compliance_var.set(f"Compliance: {pct}%")
        else:
            self.compliance_var.set(f"Compliance: 0%")
        
        text = self.output_box.get(1.0, tk.END).encode("utf-8")
        h = hashlib.sha256(text).hexdigest().upper()[:16]
        self.hash_var.set(f"Report ID: {h}")

    def show_dashboard(self):
        dashboard = tk.Toplevel(self.root)
        dashboard.title("Hardening Dashboard")
        dashboard.geometry("600x500")
        dashboard.configure(bg=BG_COLOR)
        
        header = tk.Frame(dashboard, bg=HEADER_COLOR, height=50)
        header.pack(fill=tk.X)
        tk.Label(header, text="HARDENING DASHBOARD", fg="white", bg=HEADER_COLOR,
                font=("Segoe UI", 14, "bold")).pack(pady=10)
        
        content = tk.Frame(dashboard, bg=BG_COLOR, padx=20, pady=20)
        content.pack(fill=tk.BOTH, expand=True)
        
        if self.conn:
            try:
                cursor = self.conn.cursor()
                
                # Get unique policy count (latest results only)
                cursor.execute("""
                    SELECT COUNT(DISTINCT s1.policy_id || s1.module_name) as unique_policies
                    FROM scan_results s1
                    INNER JOIN (
                        SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                        FROM scan_results 
                        GROUP BY policy_id, module_name
                    ) s2 ON s1.policy_id = s2.policy_id 
                        AND s1.module_name = s2.module_name 
                        AND s1.scan_timestamp = s2.max_timestamp
                """)
                total_policies = cursor.fetchone()[0] or 0
                
                cursor.execute("""
                    SELECT COUNT(*) as passed
                    FROM scan_results s1
                    INNER JOIN (
                        SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                        FROM scan_results 
                        GROUP BY policy_id, module_name
                    ) s2 ON s1.policy_id = s2.policy_id 
                        AND s1.module_name = s2.module_name 
                        AND s1.scan_timestamp = s2.max_timestamp
                    WHERE s1.status='PASS'
                """)
                passed = cursor.fetchone()['passed'] or 0
                
                compliance = int((passed / total_policies * 100)) if total_policies > 0 else 0
                
                # Get module counts from latest results only
                cursor.execute("""
                    SELECT s1.module_name, COUNT(*) as count
                    FROM scan_results s1
                    INNER JOIN (
                        SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                        FROM scan_results 
                        GROUP BY policy_id, module_name
                    ) s2 ON s1.policy_id = s2.policy_id 
                        AND s1.module_name = s2.module_name 
                        AND s1.scan_timestamp = s2.max_timestamp
                    GROUP BY s1.module_name
                    ORDER BY s1.module_name
                """)
                module_counts = cursor.fetchall()
                
                stats_text = f"""
                OVERALL STATISTICS
                {"="*40}
                Total Modules:     {len(MODULES)}
                Total Policies:    {total_policies}
                Passed Policies:   {passed}
                Failed Policies:   {total_policies - passed}
                Overall Compliance: {compliance}%
                
                MODULE BREAKDOWN:
                {"="*40}
                """
                
                for module, count in module_counts:
                    cursor.execute("""
                        SELECT COUNT(*) as passed
                        FROM scan_results s1
                        INNER JOIN (
                            SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                            FROM scan_results 
                            WHERE module_name=?
                            GROUP BY policy_id, module_name
                        ) s2 ON s1.policy_id = s2.policy_id 
                            AND s1.module_name = s2.module_name 
                            AND s1.scan_timestamp = s2.max_timestamp
                        WHERE s1.module_name=? AND s1.status='PASS'
                    """, (module, module))
                    module_passed = cursor.fetchone()['passed'] or 0
                    module_compliance = int((module_passed / count * 100)) if count > 0 else 0
                    stats_text += f"{module:25} {count:3} policies, {module_compliance:3}% compliance\n"
                
                stats_label = tk.Label(content, text=stats_text, bg=BG_COLOR, fg=TEXT_COLOR,
                                      font=("Consolas", 10), justify=tk.LEFT)
                stats_label.pack(anchor="w")
                
            except Exception as e:
                error_label = tk.Label(content, text=f"Error loading dashboard: {e}", 
                                      bg=BG_COLOR, fg=COLOR_FAIL)
                error_label.pack()
        else:
            error_label = tk.Label(content, text="Database not connected", 
                                  bg=BG_COLOR, fg=COLOR_FAIL)
            error_label.pack()

    def show_about(self):
        about_text = f"""Enterprise Linux Hardening Tool v3.0

Features:
• 9 Hardening Modules with Rollback
• Real-time Console Output
• Progress Tracking
• Excel Export (XLSX)
• PDF Report Generation
• Blockchain Tamper-Proofing
• Search and Filter
• Fix History Tracking
• Compliance Dashboard
• Policy Selection with Checkboxes
• Toggle between Selected/All Policies

Modules: {', '.join([m[0] for m in MODULES])}

Database: {DB_FILE}
Created: {datetime.datetime.now().strftime('%Y-%m-%d')}
        """
        messagebox.showinfo("About", about_text)

    def cleanup_duplicates(self):
        """Clean up duplicate records in the database"""
        if not self.conn:
            messagebox.showerror("Error", "Database not connected")
            return
        
        response = messagebox.askyesno("Clean Database", 
            "This will remove duplicate scan results, keeping only the latest entry for each policy.\n\n"
            "Note: This operation cannot be undone.\n\n"
            "Do you want to proceed?")
        
        if not response:
            return
        
        try:
            self.show_progress("Cleaning up database duplicates...")
            
            cursor = self.conn.cursor()
            
            # Create a temporary table with unique records
            cursor.execute('''
                CREATE TEMPORARY TABLE temp_scan_results AS
                SELECT 
                    MIN(id) as id,
                    policy_id,
                    policy_name,
                    expected_value,
                    current_value,
                    status,
                    module_name,
                    MAX(scan_timestamp) as scan_timestamp
                FROM scan_results
                GROUP BY policy_id, module_name
            ''')
            
            # Count before cleanup
            cursor.execute("SELECT COUNT(*) as count FROM scan_results")
            before_count = cursor.fetchone()['count']
            
            # Delete all records
            cursor.execute('DELETE FROM scan_results')
            
            # Insert unique records back
            cursor.execute('''
                INSERT INTO scan_results 
                (id, policy_id, policy_name, expected_value, current_value, status, module_name, scan_timestamp)
                SELECT id, policy_id, policy_name, expected_value, current_value, status, module_name, scan_timestamp
                FROM temp_scan_results
            ''')
            
            # Drop temporary table
            cursor.execute('DROP TABLE temp_scan_results')
            
            # Count after cleanup
            cursor.execute("SELECT COUNT(*) as count FROM scan_results")
            after_count = cursor.fetchone()['count']
            
            self.conn.commit()
            self.hide_progress()
            
            removed = before_count - after_count
            messagebox.showinfo("Cleanup Complete", 
                f"Database cleanup completed successfully!\n\n"
                f"Records before: {before_count}\n"
                f"Records after: {after_count}\n"
                f"Duplicates removed: {removed}")
            
            # Refresh the current view
            if self.current_module_index < len(MODULES):
                module_name = MODULES[self.current_module_index][0]
                self.load_scan_results(module_name)
                self.load_fix_history()
            
        except Exception as e:
            self.hide_progress()
            messagebox.showerror("Cleanup Error", f"Failed to clean up database:\n{str(e)}")

def main():
    root = tk.Tk()
    app = HardeningApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
