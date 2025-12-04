#!/usr/bin/env python3
import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog, messagebox, LabelFrame
import sqlite3
import subprocess
import threading
import datetime
import hashlib
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# -------------------------
# CONFIG: Modules & scripts
# -------------------------
MODULES = [
    ("Access Control", "access_control.sh"),
    ("Package Management", "package_mgmt.sh"),
    ("Filesystem", "filesystem.sh"),
    ("Services", "services.sh"),
    ("System Maintenance", "system_maintenance.sh"),
    ("Firewall", "firewall.sh"),
    ("Network", "network.sh"),
    ("User Accounts", "user_accounts.sh"),
]

ROLLBACK_SCRIPTS = {
    "Access Control": "access_control_rollback.sh",
    "Package Management": "package_mgmt_rollback.sh",
    "Filesystem": "filesystem_rollback.sh",
    "Services": "services_rollback.sh",
    "System Maintenance": "system_maintenance_rollback.sh",
    "Firewall": "firewall_rollback.sh",
    "Network": "network_rollback.sh",
    "User Accounts": "user_accounts_rollback.sh",
}

# Colors - Blue Theme
BG_COLOR = "#f0f8ff"  # Alice Blue
HEADER_COLOR = "#1e88e5"  # Bright Blue
SIDEBAR_COLOR = "#bbdefb"  # Light Blue
BUTTON_COLOR = "#2196f3"  # Material Blue
BUTTON_HOVER = "#1976d2"  # Darker Blue
TEXT_COLOR = "#212121"  # Dark Gray
COLOR_PASS = "#4caf50"  # Green
COLOR_FAIL = "#f44336"  # Red
COLOR_WARN = "#ff9800"  # Orange
COLOR_INFO = "#2196f3"  # Blue
COLOR_FIXED = "#673ab7"  # Purple
COLOR_MANUAL = "#ff9800"  # Orange
COLOR_RUNNING = "#9c27b0"  # Purple
COLOR_NORMAL = "#212121"  # Dark Gray

DB_FILE = "/home/kali/hardening.db"

# -------------------------
# Helper: Run subprocess streaming
# -------------------------
def run_command_stream(cmd, line_callback, done_callback=None, progress_callback=None):
    try:
        proc = subprocess.Popen(cmd,
                                stdout=subprocess.PIPE,
                                stderr=subprocess.STDOUT,
                                bufsize=1,
                                universal_newlines=True,
                                text=True)
    except Exception as e:
        line_callback(f"[ERROR] Failed to start: {e}")
        if done_callback:
            done_callback(-1)
        return

    lines_processed = 0
    
    try:
        for line in proc.stdout:
            if not line:
                continue
            stripped_line = line.rstrip("\n")
            line_callback(stripped_line)
            lines_processed += 1
            if progress_callback:
                progress_callback(lines_processed)
    except Exception as e:
        line_callback(f"[ERROR] Exception while reading output: {e}")
    proc.wait()
    if done_callback:
        done_callback(proc.returncode)

# -------------------------
# Main Application
# -------------------------
class HardeningApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Enterprise Linux Hardening Tool v2.0")
        self.root.geometry("1400x850")
        self.root.configure(bg=BG_COLOR)
        self.current_module_index = 0
        self.running_proc = False
        self.current_action = None
        self.filter_text = ""

        # Initialize database connection
        try:
            self.conn = sqlite3.connect(DB_FILE, check_same_thread=False)
            self.conn.row_factory = sqlite3.Row
        except Exception as e:
            messagebox.showerror("Database Error", f"Cannot connect to database:\n{DB_FILE}\nError: {e}")
            self.conn = None

        # -------------------------
        # Left Panel: Module List
        # -------------------------
        self.left_frame = tk.Frame(root, bg=SIDEBAR_COLOR, width=260, relief=tk.RAISED, borderwidth=2)
        self.left_frame.pack(side=tk.LEFT, fill=tk.Y)
        self.left_frame.pack_propagate(False)
        
        # Header
        header_left = tk.Frame(self.left_frame, bg=HEADER_COLOR, height=60)
        header_left.pack(fill=tk.X)
        tk.Label(header_left, text="HARDENING", fg="white", bg=HEADER_COLOR, 
                font=("Segoe UI", 16, "bold")).pack(side=tk.LEFT, padx=15, pady=10)
        
        # Modules Section
        modules_frame = LabelFrame(self.left_frame, text=" MODULES ", font=("Segoe UI", 11, "bold"),
                                  bg=SIDEBAR_COLOR, fg=TEXT_COLOR, relief=tk.FLAT, borderwidth=0)
        modules_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
        
        self.module_buttons = []
        for idx, (title, _) in enumerate(MODULES):
            btn_frame = tk.Frame(modules_frame, bg=SIDEBAR_COLOR)
            btn_frame.pack(fill=tk.X, padx=5, pady=2)
            
            b = tk.Button(btn_frame, text=title, width=22, height=1,
                         font=("Segoe UI", 9),
                         bg=SIDEBAR_COLOR, fg=TEXT_COLOR,
                         relief=tk.FLAT, anchor="w",
                         command=lambda i=idx: self.select_module(i))
            b.pack(side=tk.LEFT, padx=5)
            
            # Add hover effect
            b.bind("<Enter>", lambda e, b=b: b.configure(bg="#90caf9"))
            b.bind("<Leave>", lambda e, b=b: b.configure(bg=SIDEBAR_COLOR))
            
            self.module_buttons.append(b)

        # Actions Section
        actions_frame = LabelFrame(self.left_frame, text=" ACTIONS ", font=("Segoe UI", 11, "bold"),
                                  bg=SIDEBAR_COLOR, fg=TEXT_COLOR, relief=tk.FLAT, borderwidth=0)
        actions_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
        
        action_buttons = [
            ("Scan All", "#4caf50", lambda: self.start_action_all("scan")),
            ("Fix All", "#2196f3", lambda: self.start_action_all("fix")),
            ("Rollback All", "#f44336", self.rollback_all_modules),
        ]
        
        for text, color, command in action_buttons:
            btn = tk.Button(actions_frame, text=text, width=20, height=1,
                          font=("Segoe UI", 9, "bold"),
                          bg=color, fg="white",
                          relief=tk.RAISED, borderwidth=1,
                          command=command)
            btn.pack(padx=10, pady=3)
            btn.bind("<Enter>", lambda e, b=btn: b.configure(relief=tk.SUNKEN))
            btn.bind("<Leave>", lambda e, b=btn: b.configure(relief=tk.RAISED))

        # Stats Section
        stats_frame = LabelFrame(self.left_frame, text=" SYSTEM INFO ", font=("Segoe UI", 11, "bold"),
                                bg=SIDEBAR_COLOR, fg=TEXT_COLOR, relief=tk.FLAT, borderwidth=0)
        stats_frame.pack(fill=tk.X, padx=10, pady=(20, 10))
        
        self.stats_vars = {
            "Total Modules": tk.StringVar(value="8"),
            "Total Policies": tk.StringVar(value="0"),
            "Last Scan": tk.StringVar(value="Never"),
            "Database": tk.StringVar(value="Connected" if self.conn else "Error"),
        }
        
        for label, var in self.stats_vars.items():
            stat_frame = tk.Frame(stats_frame, bg=SIDEBAR_COLOR)
            stat_frame.pack(fill=tk.X, padx=5, pady=2)
            tk.Label(stat_frame, text=label+":", bg=SIDEBAR_COLOR, fg=TEXT_COLOR,
                    font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=5)
            tk.Label(stat_frame, textvariable=var, bg=SIDEBAR_COLOR, fg="#1e88e5",
                    font=("Segoe UI", 9, "bold")).pack(side=tk.RIGHT, padx=5)

        # -------------------------
        # Right Panel: Main
        # -------------------------
        self.main_frame = tk.Frame(root, bg=BG_COLOR)
        self.main_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Header
        header_frame = tk.Frame(self.main_frame, bg=HEADER_COLOR, height=70)
        header_frame.pack(fill=tk.X)
        
        title_frame = tk.Frame(header_frame, bg=HEADER_COLOR)
        title_frame.pack(side=tk.LEFT, padx=20, pady=10)
        
        tk.Label(title_frame, text="Enterprise Linux", fg="white", bg=HEADER_COLOR,
                font=("Segoe UI", 12)).pack(anchor="w")
        tk.Label(title_frame, text="HARDENING TOOL", fg="white", bg=HEADER_COLOR,
                font=("Segoe UI", 18, "bold")).pack(anchor="w")
        
        # Current module display
        self.current_module_var = tk.StringVar(value="Select a Module")
        current_module_label = tk.Label(header_frame, textvariable=self.current_module_var,
                                       fg="white", bg=HEADER_COLOR, font=("Segoe UI", 11, "bold"))
        current_module_label.pack(side=tk.RIGHT, padx=20, pady=10)

        # -------------------------
        # Control Panel with Progress Bar
        # -------------------------
        control_frame = tk.Frame(self.main_frame, bg=BG_COLOR)
        control_frame.pack(fill=tk.X, padx=20, pady=(15, 10))
        
        # Action buttons
        btn_frame = tk.Frame(control_frame, bg=BG_COLOR)
        btn_frame.pack(side=tk.LEFT)
        
        self.scan_btn = self.create_action_button(btn_frame, "Scan", "#4caf50", lambda: self.start_action("scan"))
        self.scan_btn.pack(side=tk.LEFT, padx=3)
        
        self.fix_btn = self.create_action_button(btn_frame, "Fix", "#2196f3", lambda: self.start_action("fix"))
        self.fix_btn.pack(side=tk.LEFT, padx=3)
        
        self.rollback_btn = self.create_action_button(btn_frame, "Rollback", "#f44336", self.start_module_rollback)
        self.rollback_btn.pack(side=tk.LEFT, padx=3)
        
        # Search/Filter
        search_frame = tk.Frame(control_frame, bg=BG_COLOR)
        search_frame.pack(side=tk.RIGHT)
        
        tk.Label(search_frame, text="Search:", bg=BG_COLOR, fg=TEXT_COLOR,
                font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 5))
        
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(search_frame, textvariable=self.search_var, width=25,
                                    font=("Segoe UI", 9), relief=tk.SUNKEN, borderwidth=1)
        self.search_entry.pack(side=tk.LEFT, padx=5)
        self.search_entry.bind("<KeyRelease>", self.filter_table)
        
        search_btn = tk.Button(search_frame, text="🔍", width=2, height=1,
                              bg=BUTTON_COLOR, fg="white", font=("Segoe UI", 9),
                              command=self.filter_table)
        search_btn.pack(side=tk.LEFT, padx=2)
        
        clear_btn = tk.Button(search_frame, text="Clear", width=6, height=1,
                             bg="#9e9e9e", fg="white", font=("Segoe UI", 9),
                             command=self.clear_filter)
        clear_btn.pack(side=tk.LEFT, padx=2)

        # Progress Bar
        self.progress_frame = tk.Frame(self.main_frame, bg=BG_COLOR)
        self.progress_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        
        self.progress_label = tk.Label(self.progress_frame, text="", bg=BG_COLOR, fg=TEXT_COLOR,
                                      font=("Segoe UI", 9))
        self.progress_label.pack(anchor="w")
        
        self.progress_bar = ttk.Progressbar(self.progress_frame, mode='indeterminate', length=400)
        self.progress_bar.pack(fill=tk.X, pady=(2, 0))
        
        # Initially hide progress bar
        self.progress_frame.pack_forget()

        # -------------------------
        # Tabs: Policy Table & Console
        # -------------------------
        self.tabs = ttk.Notebook(self.main_frame)
        self.tabs.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        # Style the notebook
        style = ttk.Style()
        style.configure("TNotebook", background=BG_COLOR)
        style.configure("TNotebook.Tab", background=SIDEBAR_COLOR, foreground=TEXT_COLOR)
        style.map("TNotebook.Tab", background=[("selected", HEADER_COLOR)])

        # Scan Results Tab
        self.scan_frame = tk.Frame(self.tabs, bg=BG_COLOR)
        self.tabs.add(self.scan_frame, text=" Scan Results ")
        
        # Table controls
        table_controls = tk.Frame(self.scan_frame, bg=BG_COLOR)
        table_controls.pack(fill=tk.X, padx=5, pady=(5, 0))
        
        tk.Label(table_controls, text="Scan Results Table", bg=BG_COLOR, fg=TEXT_COLOR,
                font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        
        export_btn = tk.Button(table_controls, text="Export to Excel", width=12, height=1,
                              bg="#4caf50", fg="white", font=("Segoe UI", 9),
                              command=self.export_to_excel)
        export_btn.pack(side=tk.RIGHT, padx=5)
        
        refresh_btn = tk.Button(table_controls, text="Refresh", width=8, height=1,
                               bg=BUTTON_COLOR, fg="white", font=("Segoe UI", 9),
                               command=self.refresh_table)
        refresh_btn.pack(side=tk.RIGHT, padx=5)
        
        # Create Treeview with scrollbars
        tree_frame = tk.Frame(self.scan_frame, bg=BG_COLOR)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Scrollbars
        y_scrollbar = ttk.Scrollbar(tree_frame)
        y_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        x_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        x_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Treeview
        columns = ("Policy ID", "Policy Name", "Expected", "Current", "Status")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings",
                                yscrollcommand=y_scrollbar.set,
                                xscrollcommand=x_scrollbar.set)
        
        # Configure columns
        col_widths = [120, 250, 150, 150, 100]
        for col, width in zip(columns, col_widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, minwidth=80, anchor="center")
        
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        # Configure scrollbars
        y_scrollbar.config(command=self.tree.yview)
        x_scrollbar.config(command=self.tree.xview)
        
        # Bind double-click for details
        self.tree.bind("<Double-1>", self.show_policy_details)

        # Console Tab
        self.console_frame = tk.Frame(self.tabs, bg=BG_COLOR)
        self.tabs.add(self.console_frame, text=" Console Output ")
        
        console_controls = tk.Frame(self.console_frame, bg=BG_COLOR)
        console_controls.pack(fill=tk.X, padx=5, pady=(5, 0))
        
        tk.Label(console_controls, text="Real-time Output", bg=BG_COLOR, fg=TEXT_COLOR,
                font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        
        clear_console_btn = tk.Button(console_controls, text="Clear Console", width=12, height=1,
                                     bg="#9e9e9e", fg="white", font=("Segoe UI", 9),
                                     command=self.clear_console)
        clear_console_btn.pack(side=tk.RIGHT, padx=5)
        
        save_log_btn = tk.Button(console_controls, text="Save Log", width=10, height=1,
                                bg=BUTTON_COLOR, fg="white", font=("Segoe UI", 9),
                                command=self.save_log)
        save_log_btn.pack(side=tk.RIGHT, padx=5)
        
        # Console output
        self.output_box = scrolledtext.ScrolledText(self.console_frame, wrap=tk.WORD,
                                                   font=("Consolas", 10), bg="#fafafa",
                                                   relief=tk.SUNKEN, borderwidth=1)
        self.output_box.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Configure tags for colored output
        self.output_box.tag_config("pass", foreground=COLOR_PASS)
        self.output_box.tag_config("fail", foreground=COLOR_FAIL)
        self.output_box.tag_config("warn", foreground=COLOR_WARN)
        self.output_box.tag_config("info", foreground=COLOR_INFO)
        self.output_box.tag_config("fixed", foreground=COLOR_FIXED)
        self.output_box.tag_config("manual", foreground=COLOR_MANUAL)
        self.output_box.tag_config("running", foreground=COLOR_RUNNING)
        self.output_box.tag_config("normal", foreground=COLOR_NORMAL)

        # Fix History Tab
        self.history_frame = tk.Frame(self.tabs, bg=BG_COLOR)
        self.tabs.add(self.history_frame, text=" Fix History ")
        
        history_controls = tk.Frame(self.history_frame, bg=BG_COLOR)
        history_controls.pack(fill=tk.X, padx=5, pady=(5, 0))
        
        tk.Label(history_controls, text="Fix History", bg=BG_COLOR, fg=TEXT_COLOR,
                font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        
        refresh_history_btn = tk.Button(history_controls, text="Refresh", width=8, height=1,
                                       bg=BUTTON_COLOR, fg="white", font=("Segoe UI", 9),
                                       command=self.load_fix_history)
        refresh_history_btn.pack(side=tk.RIGHT, padx=5)
        
        # History Treeview
        history_tree_frame = tk.Frame(self.history_frame, bg=BG_COLOR)
        history_tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        history_y_scrollbar = ttk.Scrollbar(history_tree_frame)
        history_y_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        history_x_scrollbar = ttk.Scrollbar(history_tree_frame, orient=tk.HORIZONTAL)
        history_x_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        history_columns = ("Policy ID", "Policy Name", "Original", "Current", "Status", "Timestamp")
        self.history_tree = ttk.Treeview(history_tree_frame, columns=history_columns, show="headings",
                                        yscrollcommand=history_y_scrollbar.set,
                                        xscrollcommand=history_x_scrollbar.set)
        
        history_col_widths = [100, 200, 120, 120, 80, 150]
        for col, width in zip(history_columns, history_col_widths):
            self.history_tree.heading(col, text=col)
            self.history_tree.column(col, width=width, minwidth=80, anchor="center")
        
        self.history_tree.pack(fill=tk.BOTH, expand=True)
        
        history_y_scrollbar.config(command=self.history_tree.yview)
        history_x_scrollbar.config(command=self.history_tree.xview)

        # -------------------------
        # Summary Panel
        # -------------------------
        summary_frame = tk.Frame(self.main_frame, bg="#e3f2fd", bd=1, relief=tk.RAISED, borderwidth=1)
        summary_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        
        # Left side: Compliance Summary
        left_sum = tk.Frame(summary_frame, bg="#e3f2fd")
        left_sum.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=15, pady=10)
        
        tk.Label(left_sum, text="COMPLIANCE SUMMARY", font=("Segoe UI", 11, "bold"),
                bg="#e3f2fd", fg=TEXT_COLOR).pack(anchor="w", pady=(0, 5))
        
        self.total_checks_var = tk.StringVar(value="Total Checks: 0")
        tk.Label(left_sum, textvariable=self.total_checks_var, bg="#e3f2fd", fg=TEXT_COLOR,
                font=("Segoe UI", 9)).pack(anchor="w")
        
        self.compliance_var = tk.StringVar(value="Compliance: 0%")
        tk.Label(left_sum, textvariable=self.compliance_var, bg="#e3f2fd", fg=TEXT_COLOR,
                font=("Segoe UI", 9)).pack(anchor="w")
        
        self.pass_var = tk.StringVar(value="Passed: 0")
        tk.Label(left_sum, textvariable=self.pass_var, bg="#e3f2fd", fg=COLOR_PASS,
                font=("Segoe UI", 9)).pack(anchor="w")
        
        self.fail_var = tk.StringVar(value="Failed: 0")
        tk.Label(left_sum, textvariable=self.fail_var, bg="#e3f2fd", fg=COLOR_FAIL,
                font=("Segoe UI", 9)).pack(anchor="w")
        
        # Right side: Quick Actions & Info
        right_sum = tk.Frame(summary_frame, bg="#e3f2fd")
        right_sum.pack(side=tk.RIGHT, padx=15, pady=10)
        
        self.hash_var = tk.StringVar(value="Report ID: Not generated")
        tk.Label(right_sum, textvariable=self.hash_var, bg="#e3f2fd", fg="#1e88e5",
                font=("Segoe UI", 8)).pack(anchor="e", pady=1)
        
        action_btn_frame = tk.Frame(right_sum, bg="#e3f2fd")
        action_btn_frame.pack(anchor="e", pady=(5, 0))
        
        tk.Button(action_btn_frame, text="Export All", width=10, height=1,
                 bg="#4caf50", fg="white", font=("Segoe UI", 9),
                 command=self.export_all_data).pack(side=tk.LEFT, padx=2)
        
        tk.Button(action_btn_frame, text="Dashboard", width=10, height=1,
                 bg=BUTTON_COLOR, fg="white", font=("Segoe UI", 9),
                 command=self.show_dashboard).pack(side=tk.LEFT, padx=2)
        
        tk.Button(action_btn_frame, text="About", width=10, height=1,
                 bg="#9e9e9e", fg="white", font=("Segoe UI", 9),
                 command=self.show_about).pack(side=tk.LEFT, padx=2)

        # -------------------------
        # Initialize
        # -------------------------
        self.reset_counters()
        if self.conn:
            self.select_module(0)
        else:
            self.append_console("[ERROR] Database not connected. Check DB_FILE path.")

    def create_action_button(self, parent, text, color, command):
        """Create a styled action button"""
        btn = tk.Button(parent, text=text, width=10, height=1,
                       font=("Segoe UI", 10, "bold"),
                       bg=color, fg="white",
                       relief=tk.RAISED, borderwidth=2,
                       command=command)
        # Hover effects
        btn.bind("<Enter>", lambda e, b=btn: b.configure(relief=tk.SUNKEN))
        btn.bind("<Leave>", lambda e, b=btn: b.configure(relief=tk.RAISED))
        return btn

    # -------------------------
    # Module & Table Management
    # -------------------------
    def select_module(self, idx):
        self.current_module_index = idx
        module_name = MODULES[idx][0]
        self.current_module_var.set(f"Module: {module_name}")
        
        # Update button states
        for i, b in enumerate(self.module_buttons):
            if i == idx:
                b.configure(bg="#90caf9", relief=tk.SUNKEN)
            else:
                b.configure(bg=SIDEBAR_COLOR, relief=tk.FLAT)
        
        if self.conn:
            self.load_scan_results(module_name)
            self.load_fix_history()
        
        self.clear_console()
        self.reset_counters()
        self.clear_filter()

    def load_scan_results(self, module_name):
        self.tree.delete(*self.tree.get_children())
        
        if not self.conn:
            self.append_console("[ERROR] Database not connected")
            return
        
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='scan_results'")
            if not cursor.fetchone():
                self.append_console(f"[INFO] scan_results table doesn't exist yet. Run a scan first.")
                return
            
            cursor.execute("""
                SELECT policy_id, policy_name, expected_value, current_value, status 
                FROM scan_results 
                WHERE module_name=?
                ORDER BY policy_id
            """, (module_name,))
            rows = cursor.fetchall()
            
            if not rows:
                self.append_console(f"[INFO] No scan results found for {module_name}. Run a scan first.")
                return
            
            # Update stats
            self.stats_vars["Total Policies"].set(str(len(rows)))
            
            # Filter rows based on search text
            filtered_rows = rows
            if self.filter_text:
                filtered_rows = [
                    r for r in rows 
                    if any(self.filter_text.lower() in str(field).lower() 
                          for field in r)
                ]
            
            for r in filtered_rows:
                status = r[4].lower() if r[4] else "normal"
                self.tree.insert("", tk.END, values=r, tags=(status,))
            
            # Update counters and summary
            self.update_counts_from_db(rows)
            self.update_summary()
            
            # Configure tag colors
            for status, color in [
                ("pass", COLOR_PASS), ("fail", COLOR_FAIL),
                ("manual", COLOR_MANUAL), ("fixed", COLOR_FIXED)
            ]:
                self.tree.tag_configure(status, foreground=color)
            
        except Exception as e:
            self.append_console(f"[ERROR] Failed to load scan results: {e}")

    def load_fix_history(self):
        self.history_tree.delete(*self.history_tree.get_children())
        
        if not self.conn or self.current_module_index >= len(MODULES):
            return
        
        module_name = MODULES[self.current_module_index][0]
        
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='fix_history'")
            if not cursor.fetchone():
                return
            
            cursor.execute("""
                SELECT policy_id, policy_name, original_value, current_value, status, fix_timestamp
                FROM fix_history 
                WHERE module_name=?
                ORDER BY fix_timestamp DESC
                LIMIT 100
            """, (module_name,))
            rows = cursor.fetchall()
            
            for r in rows:
                self.history_tree.insert("", tk.END, values=r)
                
        except Exception as e:
            self.append_console(f"[ERROR] Failed to load fix history: {e}")

    def filter_table(self, event=None):
        """Filter table based on search text"""
        self.filter_text = self.search_var.get().strip()
        if self.current_module_index < len(MODULES):
            module_name = MODULES[self.current_module_index][0]
            self.load_scan_results(module_name)

    def clear_filter(self):
        """Clear search filter"""
        self.search_var.set("")
        self.filter_text = ""
        if self.current_module_index < len(MODULES):
            module_name = MODULES[self.current_module_index][0]
            self.load_scan_results(module_name)

    def refresh_table(self):
        """Refresh the current table view"""
        if self.current_module_index < len(MODULES):
            module_name = MODULES[self.current_module_index][0]
            self.load_scan_results(module_name)
            self.load_fix_history()

    def show_policy_details(self, event):
        """Show details when a policy is double-clicked"""
        selection = self.tree.selection()
        if not selection:
            return
        
        item = self.tree.item(selection[0])
        values = item['values']
        
        if not values or len(values) < 5:
            return
        
        details_window = tk.Toplevel(self.root)
        details_window.title("Policy Details")
        details_window.geometry("500x400")
        details_window.configure(bg=BG_COLOR)
        
        # Create a scrolled text widget for details
        details_text = scrolledtext.ScrolledText(details_window, wrap=tk.WORD,
                                                font=("Consolas", 10), bg="#fafafa")
        details_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Format details
        details = f"""POLICY DETAILS
{"="*50}
Policy ID:     {values[0]}
Policy Name:   {values[1]}
Expected:      {values[2]}
Current:       {values[3]}
Status:        {values[4]}
{"="*50}
Timestamp:     {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Module:        {MODULES[self.current_module_index][0]}
        """
        
        details_text.insert(tk.END, details)
        details_text.config(state=tk.DISABLED)

    # -------------------------
    # Console Output
    # -------------------------
    def append_console(self, line):
        line = line.rstrip("\n")
        tag = "normal"
        
        if "[PASS]" in line:
            tag = "pass"
        elif "[FAIL]" in line:
            tag = "fail"
        elif "[WARN]" in line:
            tag = "warn"
        elif "[INFO]" in line:
            tag = "info"
        elif "[FIXED]" in line:
            tag = "fixed"
        elif "[MANUAL]" in line:
            tag = "manual"
        elif any(x in line for x in ["Running", "Starting", "Executing"]):
            tag = "running"
        
        self.output_box.insert(tk.END, line + "\n", tag)
        self.output_box.see(tk.END)
        
        # Update counters from console output
        if tag == "pass":
            self.count_pass += 1
            self.total_checks += 1
        elif tag == "fail":
            self.count_fail += 1
            self.total_checks += 1
        elif tag == "manual":
            self.count_manual += 1
            self.total_checks += 1
        elif tag == "fixed":
            self.count_fixed += 1
        
        self.update_summary()

    def clear_console(self):
        self.output_box.delete(1.0, tk.END)
        self.reset_counters()
        self.update_summary()

    def save_log(self):
        txt = self.output_box.get(1.0, tk.END)
        if not txt.strip():
            return
        
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        module = MODULES[self.current_module_index][0] if self.current_module_index < len(MODULES) else "unknown"
        
        path = filedialog.asksaveasfilename(
            defaultextension=".log",
            filetypes=[("Log files", "*.log"), ("Text files", "*.txt")],
            initialfile=f"hardening_{module}_{timestamp}.log"
        )
        
        if not path:
            return
        
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write(f"Hardening Log - {module}\n")
                f.write(f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("="*80 + "\n\n")
                f.write(txt)
            
            messagebox.showinfo("Log Saved", f"Log saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Save Error", f"Failed to save log:\n{e}")

    # -------------------------
    # Progress Bar Management
    # -------------------------
    def show_progress(self, message):
        self.progress_label.config(text=message)
        self.progress_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        self.progress_bar.start(10)
    
    def hide_progress(self):
        self.progress_bar.stop()
        self.progress_frame.pack_forget()
        self.progress_label.config(text="")
    
    def update_progress(self, value):
        if self.progress_bar["mode"] == "determinate":
            self.progress_bar["value"] = value
            self.root.update_idletasks()

    # -------------------------
    # Actions and Execution
    # -------------------------
    def start_action(self, action):
        if self.running_proc:
            messagebox.showwarning("Running", "Another action is running. Please wait.")
            return
        
        idx = self.current_module_index
        if idx >= len(MODULES):
            return
        
        title, fname = MODULES[idx]
        
        if not os.path.exists(fname):
            self.append_console(f"[FAIL] Script not found: {fname}")
            return
        
        if action in ["fix", "rollback"]:
            ok = messagebox.askyesno("Confirm", f"Are you sure to {action} {title}?\nThis will modify system configuration.")
            if not ok:
                return
        
        self.current_action = action
        cmd = ["sudo", "bash", fname, action]
        
        self.append_console("="*80)
        self.append_console(f"{title} - {action.upper()} - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        self.show_progress(f"Running {action} on {title}...")
        self.running_proc = True
        self._set_buttons_state("disabled")
        
        t = threading.Thread(target=lambda: self._run_and_stream(cmd, title, action))
        t.daemon = True
        t.start()

    def start_action_all(self, action):
        if self.running_proc:
            messagebox.showwarning("Running", "Another action is running. Please wait.")
            return
        
        if action in ["fix", "rollback"]:
            ok = messagebox.askyesno("Confirm", 
                f"Run {action} on ALL {len(MODULES)} modules?\nThis will modify system configuration.")
            if not ok:
                return
        
        def run_next(idx=0):
            if idx < len(MODULES):
                self.current_module_index = idx
                self.clear_console()
                self.current_module_var.set(f"Module: {MODULES[idx][0]}")
                
                # Update button state
                for i, b in enumerate(self.module_buttons):
                    if i == idx:
                        b.configure(bg="#90caf9", relief=tk.SUNKEN)
                    else:
                        b.configure(bg=SIDEBAR_COLOR, relief=tk.FLAT)
                
                # Run the action
                title, fname = MODULES[idx]
                cmd = ["sudo", "bash", fname, action]
                
                self.append_console(f"Processing: {title} ({idx+1}/{len(MODULES)})")
                self.append_console("-"*60)
                
                self.show_progress(f"Processing {title} ({idx+1}/{len(MODULES)})...")
                
                def on_line(line):
                    self.root.after(0, lambda: self.append_console(line))
                
                def on_done(returncode):
                    self.append_console(f"[INFO] Finished {title} with exit code {returncode}")
                    self.append_console("="*80)
                    self.hide_progress()
                    
                    # Load results
                    if self.conn:
                        self.root.after(100, lambda: self.load_scan_results(title))
                        self.root.after(100, lambda: self.load_fix_history())
                    
                    # Run next module
                    self.root.after(500, lambda: run_next(idx + 1))
                
                # Run in thread
                t = threading.Thread(target=lambda: run_command_stream(cmd, on_line, on_done))
                t.daemon = True
                t.start()
                
            else:
                self.append_console("[INFO] All modules completed!")
                self.hide_progress()
                self._set_buttons_state("normal")
                self.running_proc = False
                messagebox.showinfo("Complete", f"All modules {action} completed!")
                self.stats_vars["Last Scan"].set(datetime.datetime.now().strftime("%H:%M"))
        
        self.clear_console()
        self._set_buttons_state("disabled")
        self.running_proc = True
        run_next()

    def start_module_rollback(self):
        idx = self.current_module_index
        if idx >= len(MODULES):
            return
        
        title = MODULES[idx][0]
        rollback_script = ROLLBACK_SCRIPTS.get(title)
        
        if not rollback_script or not os.path.exists(rollback_script):
            self.append_console(f"[FAIL] Rollback script not found for {title}")
            messagebox.showerror("Error", f"Rollback script not found:\n{rollback_script or 'N/A'}")
            return
        
        ok = messagebox.askyesno("Confirm Rollback", 
            f"Rollback {title}?\nThis will undo all fixes made by this module.")
        if not ok:
            return
        
        cmd = ["sudo", "bash", rollback_script]
        
        self.append_console("="*80)
        self.append_console(f"ROLLBACK {title} - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        self.show_progress(f"Rolling back {title}...")
        self.running_proc = True
        self._set_buttons_state("disabled")
        
        t = threading.Thread(target=lambda: self._run_and_stream(cmd, f"{title} Rollback", "rollback"))
        t.daemon = True
        t.start()

    def rollback_all_modules(self):
        if self.running_proc:
            messagebox.showwarning("Running", "Another action is running. Please wait.")
            return
        
        ok = messagebox.askyesno("Confirm Full Rollback", 
            f"Rollback ALL {len(MODULES)} modules?\nThis will undo ALL fixes across the system!")
        if not ok:
            return
        
        def run_next(idx=0):
            if idx < len(MODULES):
                title = MODULES[idx][0]
                rollback_script = ROLLBACK_SCRIPTS.get(title)
                
                if rollback_script and os.path.exists(rollback_script):
                    self.append_console(f"Rolling back: {title} ({idx+1}/{len(MODULES)})")
                    self.show_progress(f"Rolling back {title} ({idx+1}/{len(MODULES)})...")
                    
                    cmd = ["sudo", "bash", rollback_script]
                    
                    def on_line(line):
                        self.root.after(0, lambda: self.append_console(line))
                    
                    def on_done(returncode):
                        self.append_console(f"[INFO] Finished rollback {title} with exit code {returncode}")
                        self.append_console("-"*60)
                        self.hide_progress()
                        self.root.after(100, lambda: run_next(idx + 1))
                    
                    t = threading.Thread(target=lambda: run_command_stream(cmd, on_line, on_done))
                    t.daemon = True
                    t.start()
                else:
                    self.append_console(f"[WARN] Rollback script not found for {title}")
                    self.root.after(100, lambda: run_next(idx + 1))
            
            else:
                self.append_console("[INFO] All rollbacks completed!")
                self.hide_progress()
                self._set_buttons_state("normal")
                self.running_proc = False
                messagebox.showinfo("Complete", "All rollbacks completed!")
        
        self.clear_console()
        self._set_buttons_state("disabled")
        self.running_proc = True
        run_next()

    def _run_and_stream(self, cmd, title, action):
        def on_line(line):
            self.root.after(0, lambda: self.append_console(line))
        
        def on_done(returncode):
            self.running_proc = False
            self.hide_progress()
            self._set_buttons_state("normal")
            
            if returncode == 0:
                status = "SUCCESS"
            else:
                status = "FAILED"
            
            self.append_console(f"[INFO] {title} finished with exit code {returncode} ({status})")
            
            # Refresh data
            if self.conn and self.current_module_index < len(MODULES):
                module_name = MODULES[self.current_module_index][0]
                self.root.after(500, lambda: self.load_scan_results(module_name))
                self.root.after(500, lambda: self.load_fix_history())
            
            # Update last scan time
            if action == "scan":
                self.stats_vars["Last Scan"].set(datetime.datetime.now().strftime("%H:%M"))
        
        run_command_stream(cmd, on_line, on_done)

    def _set_buttons_state(self, state):
        for w in [self.scan_btn, self.fix_btn, self.rollback_btn]:
            w.configure(state=state)

    # -------------------------
    # Export Functions
    # -------------------------
    def export_to_excel(self):
        if not self.conn:
            messagebox.showerror("Database Error", "Database not connected.")
            return
        
        module_name = MODULES[self.current_module_index][0] if self.current_module_index < len(MODULES) else "Unknown"
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"hardening_{module_name}_{timestamp}.xlsx"
        )
        
        if not path:
            return
        
        try:
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Scan Results"
            
            # Write headers
            headers = ["Policy ID", "Policy Name", "Expected Value", "Current Value", "Status", "Module", "Timestamp"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")
            
            # Get data from database
            cursor = self.conn.cursor()
            cursor.execute("""
                SELECT policy_id, policy_name, expected_value, current_value, status, module_name, scan_timestamp
                FROM scan_results 
                WHERE module_name=?
                ORDER BY policy_id
            """, (module_name,))
            rows = cursor.fetchall()
            
            # Write data
            for row_num, row in enumerate(rows, 2):
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
                    
                    # Color code based on status
                    if col_num == 5:  # Status column
                        if row[4] == "PASS":
                            ws.cell(row=row_num, column=col_num).fill = PatternFill(
                                start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
                        elif row[4] == "FAIL":
                            ws.cell(row=row_num, column=col_num).fill = PatternFill(
                                start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
                        elif row[4] == "MANUAL":
                            ws.cell(row=row_num, column=col_num).fill = PatternFill(
                                start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(path)
            
            messagebox.showinfo("Export Successful", 
                f"Exported {len(rows)} records to:\n{path}\n\nModule: {module_name}")
            
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export to Excel:\n{e}")

    def export_all_data(self):
        if not self.conn:
            messagebox.showerror("Database Error", "Database not connected.")
            return
        
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"hardening_complete_{timestamp}.xlsx"
        )
        
        if not path:
            return
        
        try:
            wb = Workbook()
            
            # Export scan results for all modules
            for module_name, _ in MODULES:
                if not wb.sheetnames:
                    ws = wb.active
                    ws.title = module_name[:31]  # Excel sheet name limit
                else:
                    ws = wb.create_sheet(title=module_name[:31])
                
                # Headers
                headers = ["Policy ID", "Policy Name", "Expected", "Current", "Status", "Timestamp"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")
                
                # Data
                cursor = self.conn.cursor()
                cursor.execute("""
                    SELECT policy_id, policy_name, expected_value, current_value, status, scan_timestamp
                    FROM scan_results 
                    WHERE module_name=?
                    ORDER BY policy_id
                """, (module_name,))
                rows = cursor.fetchall()
                
                for row_num, row in enumerate(rows, 2):
                    for col_num, value in enumerate(row, 1):
                        ws.cell(row=row_num, column=col_num, value=value)
            
            # Remove default sheet if empty
            if len(wb.sheetnames) > 1 and wb["Sheet"].max_row == 1:
                std = wb["Sheet"]
                wb.remove(std)
            
            wb.save(path)
            messagebox.showinfo("Export Successful", f"All data exported to:\n{path}")
            
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export all data:\n{e}")

    # -------------------------
    # Dashboard and Summary
    # -------------------------
    def update_counts_from_db(self, rows):
        """Update counters from database rows"""
        self.count_pass = sum(1 for r in rows if r[4] == "PASS")
        self.count_fail = sum(1 for r in rows if r[4] == "FAIL")
        self.count_manual = sum(1 for r in rows if r[4] == "MANUAL")
        self.count_fixed = sum(1 for r in rows if r[4] == "FIXED")
        self.total_checks = len(rows)

    def reset_counters(self):
        self.count_pass = 0
        self.count_fail = 0
        self.count_warn = 0
        self.count_info = 0
        self.count_fixed = 0
        self.count_manual = 0
        self.total_checks = 0

    def update_summary(self):
        self.total_checks_var.set(f"Total Checks: {self.total_checks}")
        self.pass_var.set(f"Passed: {self.count_pass}")
        self.fail_var.set(f"Failed: {self.count_fail}")
        
        total_for_compliance = self.count_pass + self.count_fail
        if total_for_compliance > 0:
            pct = int((self.count_pass / total_for_compliance) * 100)
            self.compliance_var.set(f"Compliance: {pct}%")
        else:
            self.compliance_var.set(f"Compliance: 0%")
        
        # Generate report hash
        text = self.output_box.get(1.0, tk.END).encode("utf-8")
        h = hashlib.sha256(text).hexdigest().upper()[:16]
        self.hash_var.set(f"Report ID: {h}")

    def show_dashboard(self):
        """Show a simple dashboard with statistics"""
        dashboard = tk.Toplevel(self.root)
        dashboard.title("Hardening Dashboard")
        dashboard.geometry("600x500")
        dashboard.configure(bg=BG_COLOR)
        
        # Header
        header = tk.Frame(dashboard, bg=HEADER_COLOR, height=50)
        header.pack(fill=tk.X)
        tk.Label(header, text="HARDENING DASHBOARD", fg="white", bg=HEADER_COLOR,
                font=("Segoe UI", 14, "bold")).pack(pady=10)
        
        # Content
        content = tk.Frame(dashboard, bg=BG_COLOR, padx=20, pady=20)
        content.pack(fill=tk.BOTH, expand=True)
        
        # Get overall statistics
        if self.conn:
            try:
                cursor = self.conn.cursor()
                
                # Total policies
                cursor.execute("SELECT COUNT(*) FROM scan_results")
                total_policies = cursor.fetchone()[0]
                
                # Compliance percentage
                cursor.execute("SELECT COUNT(*) FROM scan_results WHERE status='PASS'")
                passed = cursor.fetchone()[0]
                
                compliance = int((passed / total_policies * 100)) if total_policies > 0 else 0
                
                # Module counts
                cursor.execute("SELECT module_name, COUNT(*) FROM scan_results GROUP BY module_name")
                module_counts = cursor.fetchall()
                
                # Display stats
                stats_text = f"""
                OVERALL STATISTICS
                {"="*40}
                Total Modules:     {len(MODULES)}
                Total Policies:    {total_policies}
                Passed Policies:   {passed}
                Failed Policies:   {total_policies - passed}
                Overall Compliance: {compliance}%
                
                MODULE BREAKDOWN:
                {"="*40}
                """
                
                for module, count in module_counts:
                    cursor.execute("SELECT COUNT(*) FROM scan_results WHERE module_name=? AND status='PASS'", (module,))
                    module_passed = cursor.fetchone()[0]
                    module_compliance = int((module_passed / count * 100)) if count > 0 else 0
                    stats_text += f"{module:25} {count:3} policies, {module_compliance:3}% compliance\n"
                
                stats_label = tk.Label(content, text=stats_text, bg=BG_COLOR, fg=TEXT_COLOR,
                                      font=("Consolas", 10), justify=tk.LEFT)
                stats_label.pack(anchor="w")
                
            except Exception as e:
                error_label = tk.Label(content, text=f"Error loading dashboard: {e}", 
                                      bg=BG_COLOR, fg=COLOR_FAIL)
                error_label.pack()
        else:
            error_label = tk.Label(content, text="Database not connected", 
                                  bg=BG_COLOR, fg=COLOR_FAIL)
            error_label.pack()

    # -------------------------
    # About Dialog
    # -------------------------
    def show_about(self):
        about_text = f"""Enterprise Linux Hardening Tool v3.0

Features:
• 8 Hardening Modules with Rollback
• Real-time Console Output
• Progress Tracking
• Excel Export (XLSX)
• Search and Filter
• Fix History Tracking
• Compliance Dashboard

Modules: {', '.join([m[0] for m in MODULES])}

Database: {DB_FILE}
Created: {datetime.datetime.now().strftime('%Y-%m-%d')}
        """
        messagebox.showinfo("About", about_text)

# -------------------------
# Run Application
# -------------------------
def main():
    root = tk.Tk()
    app = HardeningApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
